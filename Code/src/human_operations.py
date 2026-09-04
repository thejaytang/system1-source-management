#!/usr/bin/env python3
"""Human-operation gateway for the Requirement Source Registry.

Humans decide in ``Human Operation Desktop``. This module validates those
decisions, applies controlled changes to ``Source Register``, records program
timestamps, and hides completed rows without deleting the item-level history.
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import random
import re
import shutil
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Protection
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import source_updater as updater
import manual_intake


HUMAN_SHEET = "Human Operation Desktop"
HUMAN_TABLE = "HumanOperationsTable"
HUMAN_HEADER_ROW = 2
HUMAN_DATA_START_ROW = 3
HUMAN_HEADERS = (
    "operation_id", "operation_type", "source_id", "source_title", "trigger",
    "proposed_action", "decision", "operator", "checked_at", "operator_note",
    "program_status", "program_operated_at", "program_note", "candidate_origin",
    "intake_file", "official_url", "retrieval_url", "folder_code", "issuer",
    "jurisdiction", "document_type", "source_family", "requirement_role",
    "authoritative_language", "file_format", "version", "effective_date",
    "inclusion_rationale", "authority_quality", "scope_relevance",
    "version_currency", "traceability", "access_permission",
    "automation_readiness", "operator_selection_decision", "acquisition_channel", "provenance_status",
    "update_model", "applicability_reference", "source_status", "primary_source_id",
    "source_notes", "created_at", "payload_json",
)
FINAL_STATUSES = {"APPLIED", "REJECTED"}
OPEN_STATUSES = {"PENDING", "WAITING_FOR_HUMAN", "NEEDS_REPLAN"}
PRIORITY_STATUS_ORDER = {
    "NEEDS_REPLAN": 0,
    "WAITING_FOR_HUMAN": 1,
    "PENDING": 2,
}
PRIORITY_ITEM_RANGE = "A36:F40"
PRIORITY_ITEM_LIMIT = 5
DECISIONS = {"PENDING", "ACCEPT", "REJECT", "CORRECT", "INCORRECT"}
DISPLAY_HEADER_ALIASES = {
    "task_type": "operation_type",
    "issue_summary": "trigger",
    "requested_action": "proposed_action",
    "operator_action": "decision",
    "program_result": "program_note",
}
ACTION_NORMALISATION = {"APPLY": "ACCEPT", "RETURN": "REJECT"}
TASK_ACTIONS = {
    "SOURCE_REVIEW": ["PENDING", "APPLY", "RETURN"],
    "SELECTION_REVIEW": ["PENDING", "APPLY", "RETURN"],
    "NEW_SOURCE_CANDIDATE": ["PENDING", "ACCEPT", "REJECT"],
    "MANUAL_FILE_REPLACEMENT": ["PENDING", "ACCEPT", "REJECT"],
    "RANDOM_QA_CHECK": ["PENDING", "CORRECT", "INCORRECT"],
}
SOURCE_UPDATE_FIELDS = {
    "source_title", "official_url", "retrieval_url", "folder_code", "file_format",
    "issuer", "jurisdiction", "authoritative_language", "version", "effective_date",
    "source_family", "document_type", "requirement_role", "inclusion_rationale",
    "authority_quality", "scope_relevance", "version_currency", "traceability",
    "access_permission", "automation_readiness", "source_status", "primary_source_id",
    "notes", "acquisition_channel", "provenance_status", "update_model",
    "applicability_reference", "operator_selection_decision",
}


class HumanOperationError(RuntimeError):
    pass


def human_sheet_name(config: dict[str, Any]) -> str:
    return str(config.get("human_operation_sheet") or HUMAN_SHEET)


def operation_headers(ws, require_all: bool = True) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[HUMAN_HEADER_ROW]:
        label = str(cell.value or "").strip()
        if not label:
            continue
        headers[DISPLAY_HEADER_ALIASES.get(label, label)] = cell.column
    missing = sorted(set(HUMAN_HEADERS) - headers.keys())
    if require_all and missing:
        raise HumanOperationError(f"{ws.title} is missing fields: {', '.join(missing)}")
    return headers


def operation_record(ws, row: int, headers: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row, column).value for name, column in headers.items()}


def set_operation_fields(ws, row: int, headers: dict[str, int], updates: dict[str, Any]) -> None:
    for name, value in updates.items():
        if name not in headers:
            raise HumanOperationError(f"Unknown human-operation field: {name}")
        ws.cell(row, headers[name], value)


def last_operation_row(ws, headers: dict[str, int]) -> int:
    id_col = headers["operation_id"]
    for row in range(ws.max_row, HUMAN_DATA_START_ROW - 1, -1):
        if str(ws.cell(row, id_col).value or "").strip():
            return row
    return HUMAN_HEADER_ROW


def update_operation_table(ws, final_row: int) -> None:
    if HUMAN_TABLE in ws.tables:
        final_column = get_column_letter(len(HUMAN_HEADERS))
        ws.tables[HUMAN_TABLE].ref = f"A2:{final_column}{max(HUMAN_DATA_START_ROW, final_row)}"


def prepare_human_workbook(wb, ws, headers: dict[str, int]) -> None:
    """Restore the compact operator view and refresh Dashboard priority items."""
    visible = {
        "source_id", "source_title", "operation_type", "trigger", "proposed_action",
        "decision", "operator", "operator_note", "program_status", "program_note",
    }
    support = {"operation_id", "checked_at", "program_operated_at", "created_at", "payload_json"}
    for name, column in headers.items():
        letter = get_column_letter(column)
        ws.column_dimensions[letter].hidden = name not in visible
        if name in support:
            for row in range(HUMAN_HEADER_ROW, max(ws.max_row, HUMAN_DATA_START_ROW) + 1):
                ws.cell(row, column).protection = Protection(locked=True)
    for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
        status = str(ws.cell(row, headers["program_status"]).value or "").strip().upper()
        ws.row_dimensions[row].hidden = status in FINAL_STATUSES
    update_operation_table(ws, last_operation_row(ws, headers))

    if "Dashboard" in wb.sheetnames:
        dashboard = wb["Dashboard"]
        for row in range(36, 41):
            for column in range(1, 7):
                dashboard.cell(row, column).value = None
        today = date.today()
        priority_items: list[tuple[int, int, str, dict[str, Any]]] = []
        for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
            record = operation_record(ws, row, headers)
            status = str(record.get("program_status") or "PENDING").strip().upper()
            if status not in OPEN_STATUSES:
                continue
            created = as_review_date(record.get("created_at"))
            days_open = max(0, (today - created).days) if created else 0
            source_id = str(record.get("source_id") or "").strip()
            priority_items.append((
                PRIORITY_STATUS_ORDER.get(status, len(PRIORITY_STATUS_ORDER)),
                -days_open,
                source_id,
                {
                    "source_id": record.get("source_id"),
                    "source_title": record.get("source_title"),
                    "task_type": record.get("operation_type"),
                    "issue_summary": record.get("trigger"),
                    "days_open": days_open,
                    "program_status": status,
                },
            ))

        priority_items.sort(key=lambda item: (item[0], item[1], item[2]))
        for output_row, (_, _, _, item) in enumerate(
            priority_items[:PRIORITY_ITEM_LIMIT], start=36
        ):
            values = (
                item["source_id"], item["source_title"], item["task_type"],
                item["issue_summary"], item["days_open"], item["program_status"],
            )
            for column, value in enumerate(values, 1):
                dashboard.cell(output_row, column).value = value
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


def append_operation(ws, headers: dict[str, int], values: dict[str, Any]) -> int:
    row = last_operation_row(ws, headers) + 1
    for name in HUMAN_HEADERS:
        ws.cell(row, headers[name], values.get(name))
    ws.row_dimensions[row].hidden = str(values.get("program_status") or "").upper() in FINAL_STATUSES
    apply_operation_validation(ws, row, headers, str(values.get("operation_type") or ""))
    update_operation_table(ws, row)
    return row


def apply_operation_validation(ws, row: int, headers: dict[str, int], operation_type: str) -> None:
    actions = TASK_ACTIONS.get(operation_type.strip().upper(), ["PENDING"])
    cell = ws.cell(row, headers["decision"])
    validation = DataValidation(type="list", formula1='"' + ",".join(actions) + '"', allow_blank=False)
    validation.error = f"Choose one of: {', '.join(actions)}"
    validation.errorTitle = "Invalid operator action"
    validation.prompt = f"Allowed actions for {operation_type}: {', '.join(actions)}"
    validation.promptTitle = "Operator action"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(cell)


def operation_index(ws, headers: dict[str, int]) -> dict[str, int]:
    result: dict[str, int] = {}
    for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
        operation_id = str(ws.cell(row, headers["operation_id"]).value or "").strip()
        if operation_id:
            result[operation_id] = row
    return result


def completed_review_history(config_path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """Return the latest completed review per source and trigger."""
    config = updater.read_config(config_path)
    wb = load_workbook(config["workbook"], data_only=False)
    ws = wb[human_sheet_name(config)]
    headers = operation_headers(ws, require_all=False)
    history: dict[tuple[str, str], dict[str, Any]] = {}
    for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
        record = operation_record(ws, row, headers)
        if str(record.get("program_status") or "").strip().upper() not in FINAL_STATUSES:
            continue
        source_id = str(record.get("source_id") or "").strip()
        trigger = str(record.get("trigger") or "").strip().upper()
        if not source_id or not trigger:
            continue
        completed_at = record.get("checked_at") or record.get("program_operated_at")
        completed_date = completed_at.date() if isinstance(completed_at, datetime) else completed_at
        payload = parse_payload(record.get("payload_json"))
        key = (source_id, trigger)
        existing = history.get(key)
        existing_date = as_review_date(existing.get("completed_at")) if existing else None
        current_date = as_review_date(completed_date)
        if existing_date and current_date and existing_date >= current_date:
            continue
        history[key] = {
            "completed_at": completed_date,
            "fingerprint": str(payload.get("review_fingerprint") or ""),
            "decision": str(record.get("decision") or ""),
        }
    wb.close()
    return history


def open_operation_summary(config_path: Path, today: date | None = None) -> dict[str, Any]:
    """Summarize every visible human task, regardless of which agent created it."""
    config = updater.read_config(config_path)
    today = today or updater.local_now(config).date()
    wb = load_workbook(config["workbook"], data_only=False)
    ws = wb[human_sheet_name(config)]
    headers = operation_headers(ws, require_all=False)
    items: list[dict[str, Any]] = []
    for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
        record = operation_record(ws, row, headers)
        status = str(record.get("program_status") or "PENDING").strip().upper()
        operation_id = str(record.get("operation_id") or "").strip()
        if not operation_id or status not in OPEN_STATUSES:
            continue
        created = as_review_date(record.get("created_at"))
        items.append({
            "operation_id": operation_id,
            "operation_type": str(record.get("operation_type") or "").strip(),
            "source_id": str(record.get("source_id") or "").strip(),
            "trigger": str(record.get("trigger") or "").strip(),
            "program_status": status,
            "days_open": max(0, (today - created).days) if created else None,
        })
    wb.close()
    return {
        "open_count": len(items),
        "oldest_days": max((item["days_open"] or 0 for item in items), default=0),
        "items": items,
    }


def as_review_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return datetime.fromisoformat(value.strip()).date()
        except ValueError:
            return None
    return None


def qa_batch_id(record: dict[str, Any]) -> str:
    payload = parse_payload(record.get("payload_json"))
    batch_id = str(payload.get("batch_id") or "").strip()
    if batch_id:
        return batch_id
    operation_id = str(record.get("operation_id") or "").strip()
    match = re.match(r"^QA-([0-9]{8}T[0-9]{6})-", operation_id)
    return match.group(1) if match else operation_id

def parse_payload(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise HumanOperationError(f"Invalid payload_json: {exc}") from exc
    if not isinstance(parsed, dict):
        raise HumanOperationError("payload_json must contain a JSON object.")
    return parsed


def append_note(existing: Any, addition: str) -> str:
    base = str(existing or "").strip()
    addition = addition.strip()
    return f"{base} {addition}".strip() if base else addition


def find_source_row(ws, headers: dict[str, int], source_id: str) -> int | None:
    source_id = source_id.strip()
    for row in range(3, ws.max_row + 1):
        if str(ws.cell(row, headers["source_id"]).value or "").strip() == source_id:
            return row
    return None


def copy_source_row_template(ws, source_row: int, target_row: int, headers: dict[str, int]) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for field in ("selection_status", "current_snapshot_date", "needs_human_action"):
        column = headers[field]
        formula = ws.cell(source_row, column).value
        if isinstance(formula, str) and formula.startswith("="):
            ws.cell(target_row, column, Translator(formula, origin=ws.cell(source_row, column).coordinate).translate_formula(ws.cell(target_row, column).coordinate))


def safe_filename_title(title: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", title).strip("_")
    return (cleaned or "source")[:120]


def next_source_id(ws, headers: dict[str, int], prefix: str) -> str:
    values: list[int] = []
    pattern = re.compile(rf"^{re.escape(prefix)}(\d{{3}})$")
    for row in range(3, ws.max_row + 1):
        match = pattern.match(str(ws.cell(row, headers["source_id"]).value or "").strip())
        if match:
            values.append(int(match.group(1)))
    next_number = max(values, default=0) + 1
    if next_number > 999:
        raise HumanOperationError(f"No source IDs remain for prefix {prefix}.")
    return f"{prefix}{next_number:03d}"


EDITABLE_CANDIDATE_FIELDS = (
    "source_title", "official_url", "retrieval_url", "folder_code", "issuer",
    "jurisdiction", "document_type", "source_family", "requirement_role",
    "authoritative_language", "file_format", "version", "effective_date",
    "inclusion_rationale", "authority_quality", "scope_relevance",
    "version_currency", "traceability", "access_permission",
    "automation_readiness", "operator_selection_decision", "acquisition_channel",
    "provenance_status", "update_model", "applicability_reference",
    "source_status", "primary_source_id", "source_notes",
)

OPERATION_SOURCE_FIELD_MAP = {
    **{field: field for field in EDITABLE_CANDIDATE_FIELDS if field not in {"source_notes"}},
    "source_notes": "notes",
}


def source_record_fingerprint(record: dict[str, Any]) -> str:
    payload = {field: str(record.get(field) or "").strip() for field in sorted(SOURCE_UPDATE_FIELDS)}
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()[:20]


def candidate_payload_from_operation(record: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    """Overlay human-completed desktop fields onto a generated candidate."""
    merged = dict(payload)
    for field in EDITABLE_CANDIDATE_FIELDS:
        value = record.get(field)
        if value not in (None, ""):
            merged[OPERATION_SOURCE_FIELD_MAP.get(field, field)] = value
    merged["candidate_origin"] = record.get("candidate_origin") or payload.get("candidate_origin") or "AGENT_DISCOVERY"
    merged["intake_file"] = record.get("intake_file") or payload.get("intake_file")
    return merged


def source_updates_from_operation(record: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    """Return the human-visible source fields that an accepted operation may apply."""
    updates = dict(payload.get("updates") or {})
    for operation_field, source_field in OPERATION_SOURCE_FIELD_MAP.items():
        value = record.get(operation_field)
        if value not in (None, ""):
            updates[source_field] = value
    invalid = sorted(set(updates) - SOURCE_UPDATE_FIELDS)
    if invalid:
        raise HumanOperationError(f"The proposed update contains non-editable fields: {', '.join(invalid)}")
    return updates


def preferred_retrieval(payload: dict[str, Any], config: dict[str, Any]) -> tuple[str, str]:
    """Choose an equivalent official HTML source before lower-ranked formats."""
    candidates: list[tuple[str, str]] = []
    for field, forced_format in (("official_html_url", "html"), ("html_url", "html"), ("retrieval_url", ""), ("pdf_url", "pdf")):
        value = str(payload.get(field) or "").strip()
        if not value:
            continue
        candidates.append((value, forced_format or manual_intake.url_format(value)))
    if not candidates:
        return "", str(payload.get("file_format") or "").strip().lower()
    rank = {str(value).lower(): index for index, value in enumerate(config.get("format_preference", ["html", "pdf", "xlsx", "zip"]))}
    candidates.sort(key=lambda item: rank.get(item[1], 99))
    return candidates[0]


def archive_intake_file(payload: dict[str, Any], config: dict[str, Any], source_id: str, bucket: str) -> Path | None:
    filename = str(payload.get("intake_file") or "").strip()
    root = config.get("manual_intake_root")
    archive_root = config.get("manual_intake_archive_root")
    if not filename or root is None or archive_root is None:
        return None
    source = (root / filename).resolve()
    if source.parent != root.resolve() or not source.is_file():
        return None
    target_dir = archive_root / bucket / source_id
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / source.name
    counter = 1
    while target.exists():
        target = target_dir / f"{source.stem}_{counter}{source.suffix}"
        counter += 1
    shutil.move(str(source), str(target))
    return target


def promote_manual_candidate(payload: dict[str, Any], config: dict[str, Any], folder_code: str, stored_filename: str) -> dict[str, Any]:
    """Validate and promote a human file only when it remains the chosen format."""
    filename = str(payload.get("intake_file") or "").strip()
    root = config.get("manual_intake_root")
    if not filename or root is None:
        return {}
    source = (root / filename).resolve()
    if source.parent != root.resolve() or not source.is_file():
        return {}
    manual_format = manual_intake.normalise_format(payload.get("manual_file_format") or source.suffix)
    selected_format = manual_intake.normalise_format(payload.get("file_format") or "")
    if manual_format != selected_format:
        return {}
    byte_count = source.stat().st_size
    if selected_format == "pdf":
        updater.validate_pdf(source, byte_count, int(config["minimum_pdf_bytes"]))
    elif selected_format == "html":
        updater.validate_html(source, byte_count, int(config["minimum_html_bytes"]))
    elif selected_format in {"xlsx", "zip"}:
        updater.validate_zip_container(source, byte_count, int(config["minimum_binary_bytes"]), require_xlsx=selected_format == "xlsx")
    destination = config["source_root"] / folder_code / stored_filename
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    now = updater.local_now(config)
    return {
        "snapshot_status": "STORED",
        "download_status": "SUCCESS",
        "last_success_at": now,
        "content_change": "INITIAL",
        "current_origin": "MANUAL",
        "manual_update_date": now.date(),
        "content_hash": updater.sha256_file(destination),
        "failure_stage": "NONE",
        "failure_reason": None,
    }


def replace_source_with_manual_file(
    source_ws,
    source_headers: dict[str, int],
    record: dict[str, Any],
    payload: dict[str, Any],
    config: dict[str, Any],
    now: datetime,
) -> str:
    """Promote an authorised intake file as the next snapshot of an existing source."""
    source_id = str(record.get("source_id") or payload.get("source_id") or "").strip()
    source_row = find_source_row(source_ws, source_headers, source_id)
    if source_row is None:
        raise HumanOperationError(f"Source not found: {source_id}")
    intake_name = str(record.get("intake_file") or payload.get("intake_file") or "").strip()
    intake_root = config.get("manual_intake_root")
    if not intake_name or intake_root is None:
        raise HumanOperationError("Manual replacement requires an intake file.")
    intake_path = (intake_root / intake_name).resolve()
    if intake_path.parent != intake_root.resolve() or not intake_path.is_file():
        raise HumanOperationError(f"Manual intake file is not available: {intake_name}")

    source = updater.record_from_row(source_ws, source_row, source_headers)
    selected_format = manual_intake.normalise_format(record.get("file_format") or payload.get("manual_file_format") or intake_path.suffix)
    actual_format = manual_intake.normalise_format(intake_path.suffix)
    if selected_format != actual_format:
        raise HumanOperationError(f"The selected file_format {selected_format} does not match the supplied {actual_format} file.")
    byte_count = intake_path.stat().st_size
    if selected_format == "pdf":
        updater.validate_pdf(intake_path, byte_count, int(config["minimum_pdf_bytes"]))
    elif selected_format == "html":
        updater.validate_html(intake_path, byte_count, int(config["minimum_html_bytes"]))
    elif selected_format in {"xlsx", "zip"}:
        updater.validate_zip_container(intake_path, byte_count, int(config["minimum_binary_bytes"]), require_xlsx=selected_format == "xlsx")
    else:
        raise HumanOperationError(f"Unsupported manual replacement format: {selected_format}")

    next_snapshot = updater.next_snapshot_id(source_id, str(source.get("snapshot_id") or f"{source_id}-000"))
    filename_record = dict(source)
    filename_record["file_format"] = selected_format
    if record.get("source_title"):
        filename_record["source_title"] = record.get("source_title")
    new_filename = updater.filename_for_snapshot(filename_record, next_snapshot)
    folder_code = str(record.get("folder_code") or source.get("folder_code") or "").strip()
    old_path = config["source_root"] / str(source.get("folder_code") or folder_code) / str(source.get("stored_filename") or "")
    new_path = config["source_root"] / folder_code / new_filename
    updater.promote_download(intake_path, new_path)
    if old_path.is_file() and old_path != new_path:
        updater.archive_current(old_path, config["source_root"], str(source.get("folder_code") or folder_code), source_id)
        old_path.unlink(missing_ok=True)

    updates = source_updates_from_operation(record, payload)
    updates.update({
        "snapshot_id": next_snapshot,
        "stored_filename": new_filename,
        "snapshot_status": "STORED",
        "download_status": "SUCCESS",
        "last_success_at": now,
        "content_change": "CHANGED" if str(source.get("snapshot_status") or "").upper() == "STORED" else "INITIAL",
        "current_origin": "MANUAL",
        "manual_update_date": now.date(),
        "manual_updated_by": str(record.get("operator") or "").strip(),
        "failure_stage": "NONE",
        "failure_reason": None,
        "content_hash": updater.sha256_file(new_path),
        "file_format": selected_format,
        "folder_code": folder_code,
    })
    for field, value in updates.items():
        if field in source_headers and field not in {"selection_status", "current_snapshot_date", "needs_human_action"}:
            source_ws.cell(source_row, source_headers[field], value)
    archive_intake_file(payload | {"intake_file": intake_name}, config, source_id, "accepted")
    return source_id


def add_candidate_to_source_register(ws, headers: dict[str, int], payload: dict[str, Any], config: dict[str, Any], now: datetime, operator: str, note: str) -> str:
    title = str(payload.get("source_title") or "").strip()
    official_url = str(payload.get("official_url") or "").strip()
    folder_code = str(payload.get("folder_code") or "").strip()
    retrieval_url, file_format = preferred_retrieval(payload, config)
    payload["retrieval_url"] = retrieval_url
    payload["file_format"] = file_format
    if not title or not official_url or not folder_code or not file_format:
        raise HumanOperationError("Accepted candidates require source_title, official_url, folder_code and file_format.")
    prefix = str(config["folder_prefix_map"].get(folder_code) or "").strip()
    if not prefix:
        raise HumanOperationError(f"No source ID prefix is configured for folder_code {folder_code}.")
    for row in range(3, ws.max_row + 1):
        if str(ws.cell(row, headers["official_url"]).value or "").strip() == official_url:
            raise HumanOperationError(f"The accepted candidate duplicates existing source {ws.cell(row, headers['source_id']).value}.")
    source_id = next_source_id(ws, headers, prefix)
    target_row = max(3, ws.max_row + 1)
    template_row = max(3, target_row - 1)
    copy_source_row_template(ws, template_row, target_row, headers)
    snapshot_id = f"{source_id}-001"
    stored_filename = f"{snapshot_id}_{safe_filename_title(title)}.{file_format}"
    defaults = {
        "source_id": source_id,
        "snapshot_id": snapshot_id,
        "source_title": title,
        "official_url": official_url,
        "retrieval_url": retrieval_url,
        "folder_code": folder_code,
        "file_format": file_format,
        "stored_filename": stored_filename,
        "snapshot_status": "NOT_COLLECTED",
        "issuer": payload.get("issuer"),
        "jurisdiction": payload.get("jurisdiction") or "Other",
        "authoritative_language": payload.get("authoritative_language") or "Other",
        "version": payload.get("version"),
        "effective_date": payload.get("effective_date"),
        "source_family": payload.get("source_family") or "Pending classification",
        "document_type": payload.get("document_type") or "Other",
        "requirement_role": payload.get("requirement_role") or "Draft working material",
        "inclusion_rationale": payload.get("inclusion_rationale") or "Accepted into the candidate register after human review; formal applicability remains pending.",
        "authority_quality": payload.get("authority_quality") or "MEDIUM",
        "scope_relevance": payload.get("scope_relevance") or "MEDIUM",
        "version_currency": payload.get("version_currency") or "MEDIUM",
        "traceability": payload.get("traceability") or "MEDIUM",
        "access_permission": payload.get("access_permission") or "HIGH",
        "automation_readiness": payload.get("automation_readiness") or "MEDIUM",
        "operator_selection_decision": payload.get("operator_selection_decision") or "PENDING",
        "download_status": "NOT_RUN",
        "last_attempt_at": None,
        "last_success_at": None,
        "content_change": "UNKNOWN",
        "current_origin": "NONE",
        "manual_update_date": None,
        "manual_updated_by": None,
        "failure_stage": "NONE",
        "failure_reason": None,
        "content_hash": None,
        "source_status": "CURRENT",
        "primary_source_id": payload.get("primary_source_id"),
        "notes": append_note(payload.get("notes"), f"Human gate accepted {now.date().isoformat()} by {operator}. {note}"),
        "acquisition_channel": payload.get("acquisition_channel") or "OFFICIAL_WEBSITE",
        "provenance_status": payload.get("provenance_status") or "UNVERIFIED",
        "update_model": payload.get("update_model") or "ROLLING",
        "applicability_reference": payload.get("applicability_reference"),
    }
    try:
        manual_state = promote_manual_candidate(payload, config, folder_code, stored_filename)
    except updater.UpdaterError as exc:
        raise HumanOperationError(f"Manual intake file validation failed: {exc}") from exc
    if manual_state:
        manual_state["manual_updated_by"] = operator
        defaults.update(manual_state)
    formula_fields = {"selection_status", "current_snapshot_date", "needs_human_action"}
    for name, value in defaults.items():
        if name in headers and name not in formula_fields:
            ws.cell(target_row, headers[name], value)
    for table in ws.tables.values():
        if table.ref.startswith("A2:"):
            table.ref = f"A2:AS{target_row}"
    archive_intake_file(payload, config, source_id, "accepted")
    return source_id


def review_operation_values(item: dict[str, Any], source: dict[str, Any], now: datetime) -> dict[str, Any]:
    action_map = {
        "DOWNLOAD_FAILURE": "RETRY_SOURCE",
        "PAYWALL_BLOCKED": "APPLY_FIELD_UPDATES",
        "MISSING_FILE": "RETRY_SOURCE",
        "ROLLING_CONTENT_CHANGED": "ACKNOWLEDGE_REVIEW",
        "PROVENANCE_UNVERIFIED": "APPLY_FIELD_UPDATES",
        "APPLICABILITY_UNCONFIRMED": "APPLY_FIELD_UPDATES",
        "RANDOM_QA_FAILURE": "APPLY_FIELD_UPDATES",
        "SELECTION_PENDING": "APPLY_FIELD_UPDATES",
        "MISSING_OFFICIAL_URL": "APPLY_FIELD_UPDATES",
        "MISSING_RETRIEVAL_URL": "APPLY_FIELD_UPDATES",
        "DUPLICATE_OFFICIAL_URL": "APPLY_FIELD_UPDATES",
        "REGISTRY_FILE_MISMATCH": "RETRY_SOURCE",
    }
    issue_codes = item.get("issue_codes") or [code.strip() for code in str(item.get("trigger") or "").split(";") if code.strip()]
    mapped_actions = {action_map.get(str(code), "APPLY_FIELD_UPDATES") for code in issue_codes}
    retry_after_apply = "RETRY_SOURCE" in mapped_actions and len(mapped_actions) > 1
    machine_action = "APPLY_FIELD_UPDATES" if retry_after_apply else next(iter(mapped_actions), "APPLY_FIELD_UPDATES")
    requested_action = str(item.get("recommended_action") or "Review the source and complete the requested fields.").strip()
    payload = {
        "action": machine_action,
        "source_id": item.get("source_id"),
        "updates": item.get("proposed_updates") or {},
        "review_fingerprint": item.get("review_fingerprint") or "",
        "source_fingerprint": source_record_fingerprint(source),
        "issue_codes": issue_codes,
        "retry_after_apply": retry_after_apply,
    }
    values = {
        "operation_id": f"REV-{item['review_id']}",
        "operation_type": "SELECTION_REVIEW" if issue_codes == ["SELECTION_PENDING"] else "SOURCE_REVIEW",
        "source_id": item.get("source_id"),
        "source_title": item.get("source_title"),
        "trigger": item.get("trigger"),
        "proposed_action": requested_action,
        "decision": "PENDING",
        "program_status": "PENDING",
        "program_note": str(item.get("reason") or "").strip(),
        "created_at": now,
        "payload_json": json.dumps(payload, ensure_ascii=False, sort_keys=True),
    }
    for operation_field, source_field in OPERATION_SOURCE_FIELD_MAP.items():
        values[operation_field] = source.get(source_field)
    proposed_updates = item.get("proposed_updates") or {}
    for operation_field, source_field in OPERATION_SOURCE_FIELD_MAP.items():
        if source_field in proposed_updates:
            values[operation_field] = proposed_updates[source_field]
    return values


def sync_review_queue(config_path: Path, review_queue: list[dict[str, Any]]) -> dict[str, Any]:
    config = updater.read_config(config_path)
    now = updater.local_now(config)
    added: list[str] = []
    superseded: list[str] = []
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        if human_sheet_name(config) not in wb.sheetnames:
            wb.close()
            raise HumanOperationError(f"Missing worksheet: {human_sheet_name(config)}")
        ops = wb[human_sheet_name(config)]
        op_headers = operation_headers(ops)
        existing = operation_index(ops, op_headers)
        source_ws = wb[config["sheet_name"]]
        source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
        sources = {}
        for row in range(int(config["data_start_row"]), source_ws.max_row + 1):
            record = updater.record_from_row(source_ws, row, source_headers)
            source_id = str(record.get("source_id") or "").strip()
            if source_id:
                sources[source_id] = record
        desired_ids = {f"REV-{item['review_id']}" for item in review_queue}
        for row in range(HUMAN_DATA_START_ROW, last_operation_row(ops, op_headers) + 1):
            record = operation_record(ops, row, op_headers)
            operation_id = str(record.get("operation_id") or "").strip()
            operation_type = str(record.get("operation_type") or "").strip().upper()
            status = str(record.get("program_status") or "PENDING").strip().upper()
            action = str(record.get("decision") or "PENDING").strip().upper()
            if operation_type not in {"SOURCE_REVIEW", "SELECTION_REVIEW"} or status not in OPEN_STATUSES:
                continue
            if operation_id in desired_ids or action != "PENDING":
                continue
            set_operation_fields(ops, row, op_headers, {
                "program_status": "APPLIED",
                "program_operated_at": now,
                "program_note": "Superseded by the current consolidated source task or resolved source state.",
            })
            ops.row_dimensions[row].hidden = True
            superseded.append(operation_id)
        for item in review_queue:
            operation_id = f"REV-{item['review_id']}"
            if operation_id in existing:
                continue
            values = review_operation_values(item, sources.get(str(item.get("source_id") or ""), {}), now)
            append_operation(ops, op_headers, values)
            added.append(operation_id)
        if added or superseded:
            prepare_human_workbook(wb, ops, op_headers)
            updater.create_backup(config["workbook"], config["backup_root"], now)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()
    return {"added_count": len(added), "operation_ids": added, "superseded_count": len(superseded), "superseded_operation_ids": superseded}


def discovery_inbox_path(config_path: Path, config: dict[str, Any]) -> Path:
    raw = str(config.get("discovery_candidate_inbox") or "../runtime/inbox/discovery_candidates.json")
    return (config_path.parent / raw).resolve()


def load_discovery_candidates(config_path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    path = discovery_inbox_path(config_path, config)
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise HumanOperationError("The discovery candidate inbox must contain a JSON list.")
    return [item for item in data if isinstance(item, dict)]


def candidate_operation_id(candidate: dict[str, Any]) -> str:
    stable = str(candidate.get("content_hash") or "").strip().lower() or (
        f"{str(candidate.get('official_url') or '').strip().lower()}|{str(candidate.get('source_title') or '').strip().lower()}"
    )
    return "DISC-" + hashlib.sha256(stable.encode("utf-8")).hexdigest()[:16]


def sync_discovery_candidates(config_path: Path) -> dict[str, Any]:
    config = updater.read_config(config_path)
    candidates = load_discovery_candidates(config_path, config)
    if not candidates:
        return {"candidate_count": 0, "added_count": 0, "operation_ids": []}
    now = updater.local_now(config)
    added: list[str] = []
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        ops = wb[human_sheet_name(config)]
        op_headers = operation_headers(ops)
        existing = operation_index(ops, op_headers)
        source_ws = wb[config["sheet_name"]]
        source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
        existing_urls = {
            str(source_ws.cell(row, source_headers["official_url"]).value or "").strip():
            str(source_ws.cell(row, source_headers["source_id"]).value or "").strip()
            for row in range(int(config["data_start_row"]), source_ws.max_row + 1)
            if str(source_ws.cell(row, source_headers["official_url"]).value or "").strip()
        }
        queued_urls = {
            str(ops.cell(row, op_headers["official_url"]).value or "").strip()
            for row in range(HUMAN_DATA_START_ROW, last_operation_row(ops, op_headers) + 1)
            if str(ops.cell(row, op_headers["program_status"]).value or "PENDING").strip().upper() in OPEN_STATUSES
        }
        for candidate in candidates:
            operation_id = candidate_operation_id(candidate)
            url = str(candidate.get("official_url") or "").strip()
            is_manual_drop = str(candidate.get("candidate_origin") or "").strip().upper() == "HUMAN_DROP"
            if not str(candidate.get("source_title") or "").strip():
                continue
            if operation_id in existing or (url and url in queued_urls):
                continue
            duplicate_source_id = existing_urls.get(url, "")
            manual_replacement = bool(is_manual_drop and duplicate_source_id)
            values = {
                "operation_id": operation_id,
                "operation_type": "MANUAL_FILE_REPLACEMENT" if manual_replacement else "NEW_SOURCE_CANDIDATE",
                "source_id": duplicate_source_id,
                "source_title": candidate.get("source_title"),
                "trigger": "HUMAN_FILE_REPLACEMENT" if manual_replacement else ("POSSIBLE_DUPLICATE" if duplicate_source_id else ("HUMAN_FILE_DROP" if is_manual_drop else "DISCOVERY_AGENT_FINDING")),
                "proposed_action": "REPLACE_CURRENT_FILE" if manual_replacement else "ADD_TO_SOURCE_REGISTER",
                "decision": "PENDING",
                "program_status": "PENDING",
                "program_note": (
                    (f"The authorised file appears to match existing source {duplicate_source_id}. Confirm the original format and accept or reject the controlled replacement."
                     if manual_replacement else
                     f"The candidate may duplicate existing source {duplicate_source_id}. Confirm whether to reject it or correct its identity before acceptance.")
                    if duplicate_source_id else
                    "A newly discovered source must be accepted by a human before it receives a formal source_id."
                ),
                "candidate_origin": candidate.get("candidate_origin") or "AGENT_DISCOVERY",
                "intake_file": candidate.get("intake_file"),
                "official_url": url,
                "retrieval_url": candidate.get("retrieval_url"),
                "folder_code": candidate.get("folder_code"),
                "issuer": candidate.get("issuer"),
                "jurisdiction": candidate.get("jurisdiction"),
                "document_type": candidate.get("document_type"),
                "source_family": candidate.get("source_family"),
                "requirement_role": candidate.get("requirement_role"),
                "authoritative_language": candidate.get("authoritative_language"),
                "file_format": candidate.get("file_format"),
                "version": candidate.get("version"),
                "effective_date": candidate.get("effective_date"),
                "inclusion_rationale": candidate.get("inclusion_rationale"),
                "authority_quality": candidate.get("authority_quality"),
                "scope_relevance": candidate.get("scope_relevance"),
                "version_currency": candidate.get("version_currency"),
                "traceability": candidate.get("traceability"),
                "access_permission": candidate.get("access_permission"),
                "automation_readiness": candidate.get("automation_readiness"),
                "operator_selection_decision": candidate.get("operator_selection_decision") or "PENDING",
                "acquisition_channel": candidate.get("acquisition_channel"),
                "provenance_status": candidate.get("provenance_status"),
                "update_model": candidate.get("update_model"),
                "applicability_reference": candidate.get("applicability_reference"),
                "source_status": candidate.get("source_status") or "CURRENT",
                "primary_source_id": candidate.get("primary_source_id"),
                "source_notes": candidate.get("notes"),
                "created_at": now,
                "payload_json": json.dumps(candidate, ensure_ascii=False, sort_keys=True),
            }
            append_operation(ops, op_headers, values)
            added.append(operation_id)
            if url:
                queued_urls.add(url)
        if added:
            prepare_human_workbook(wb, ops, op_headers)
            updater.create_backup(config["workbook"], config["backup_root"], now)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()
    return {"candidate_count": len(candidates), "added_count": len(added), "operation_ids": added}


def generate_random_qa(
    config_path: Path,
    rng: random.Random | random.SystemRandom | None = None,
    today: date | None = None,
    force: bool = False,
) -> dict[str, Any]:
    config = updater.read_config(config_path)
    qa = config.get("random_qa", {})
    rng = rng or random.SystemRandom()
    now = updater.local_now(config)
    today = today or now.date()
    scheduled_day = calendar.monthrange(today.year, today.month)[1]
    catch_up = str(qa.get("schedule") or "").strip().upper() == "LAST_DAY_OF_MONTH_WITH_CATCH_UP"
    if today.day == scheduled_day:
        target_date = today
    elif catch_up:
        target_date = today.replace(day=1) - timedelta(days=1)
    else:
        target_date = today
    if not force and (not bool(qa.get("enabled", True)) or (today.day != scheduled_day and not catch_up)):
        return {
            "status": "NOT_DUE",
            "scheduled_for": date(today.year, today.month, scheduled_day).isoformat(),
            "generated_count": 0,
            "operation_ids": [],
        }
    try:
        sample_size = int(qa.get("monthly_sample_size", 5))
    except (TypeError, ValueError):
        sample_size = 0
    if not 1 <= sample_size <= int(qa.get("max_sample_size", 20)):
        raise HumanOperationError("Monthly random QA sample size must be within the configured range.")
    batch_month = target_date.strftime("%Y-%m")
    batch_id = target_date.strftime("%Y%m")
    target_last_day = calendar.monthrange(target_date.year, target_date.month)[1]
    scheduled_for = date(target_date.year, target_date.month, target_last_day)
    generated: list[str] = []
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        ops = wb[human_sheet_name(config)]
        op_headers = operation_headers(ops)
        existing_batch = False
        open_random_ids = {
            str(ops.cell(row, op_headers["source_id"]).value or "").strip()
            for row in range(HUMAN_DATA_START_ROW, last_operation_row(ops, op_headers) + 1)
            if str(ops.cell(row, op_headers["operation_type"]).value or "").strip() == "RANDOM_QA_CHECK"
            and str(ops.cell(row, op_headers["program_status"]).value or "").strip() in OPEN_STATUSES
        }
        for row in range(HUMAN_DATA_START_ROW, last_operation_row(ops, op_headers) + 1):
            operation_type = str(ops.cell(row, op_headers["operation_type"]).value or "").strip()
            if operation_type not in {"RANDOM_QA_CHECK", "RANDOM_QA_BATCH_SUMMARY"}:
                continue
            payload = parse_payload(ops.cell(row, op_headers["payload_json"]).value)
            operation_id = str(ops.cell(row, op_headers["operation_id"]).value or "").strip()
            legacy_month = ""
            match = re.match(r"^QA-(\d{4})(\d{2})", operation_id)
            if match:
                legacy_month = f"{match.group(1)}-{match.group(2)}"
            if str(payload.get("batch_month") or legacy_month).strip() == batch_month:
                existing_batch = True
                break
        if existing_batch:
            wb.close()
            return {"status": "ALREADY_CREATED", "batch_month": batch_month, "generated_count": 0, "operation_ids": []}
        source_ws = wb[config["sheet_name"]]
        source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
        eligible: list[dict[str, Any]] = []
        for row in range(int(config["data_start_row"]), source_ws.max_row + 1):
            record = updater.record_from_row(source_ws, row, source_headers)
            source_id = str(record.get("source_id") or "").strip()
            if not source_id or source_id in open_random_ids:
                continue
            if str(record.get("source_status") or "").strip().upper() != "CURRENT":
                continue
            if updater.selection_from_scores(record) != "INCLUDE":
                continue
            eligible.append(record)
        selected = rng.sample(eligible, min(sample_size, len(eligible)))
        for record in selected:
            source_id = str(record["source_id"]).strip()
            operation_id = f"QA-{batch_id}-{source_id}"
            values = {
                "operation_id": operation_id,
                "operation_type": "RANDOM_QA_CHECK",
                "source_id": source_id,
                "source_title": record.get("source_title"),
                "trigger": "MONTH_END_RANDOM_SAMPLE",
                "proposed_action": "VERIFY_SOURCE_RECORD",
                "decision": "PENDING",
                "program_status": "PENDING",
                "program_note": "Check the source identity, URL, classification, assessment and current file. Choose CORRECT or INCORRECT.",
                "official_url": record.get("official_url"),
                "issuer": record.get("issuer"),
                "jurisdiction": record.get("jurisdiction"),
                "document_type": record.get("document_type"),
                "source_family": record.get("source_family"),
                "requirement_role": record.get("requirement_role"),
                "authoritative_language": record.get("authoritative_language"),
                "file_format": record.get("file_format"),
                "created_at": now,
                "payload_json": json.dumps({
                    "action": "VERIFY_SOURCE_RECORD",
                    "source_id": source_id,
                    "batch_id": batch_id,
                    "batch_month": batch_month,
                    "scheduled_for": scheduled_for.isoformat(),
                }, sort_keys=True),
            }
            for operation_field, source_field in OPERATION_SOURCE_FIELD_MAP.items():
                values[operation_field] = record.get(source_field)
            append_operation(ops, op_headers, values)
            generated.append(operation_id)
        if generated:
            prepare_human_workbook(wb, ops, op_headers)
            updater.create_backup(config["workbook"], config["backup_root"], now)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()
    return {
        "status": "PASS",
        "batch_month": batch_month,
        "scheduled_for": scheduled_for.isoformat(),
        "requested_count": sample_size,
        "generated_count": len(generated),
        "operation_ids": generated,
    }


def apply_source_action(source_ws, source_headers: dict[str, int], record: dict[str, Any], payload: dict[str, Any], now: datetime) -> str:
    source_id = str(record.get("source_id") or payload.get("source_id") or "").strip()
    row = find_source_row(source_ws, source_headers, source_id)
    if row is None:
        raise HumanOperationError(f"Source not found: {source_id}")
    action = str(payload.get("action") or record.get("proposed_action") or "").strip().upper()
    operator = str(record.get("operator") or "").strip()
    note = str(record.get("operator_note") or "").strip()
    if action == "WITHDRAW_SOURCE":
        source_ws.cell(row, source_headers["scope_relevance"], "LOW")
        source_ws.cell(row, source_headers["source_status"], "WITHDRAWN")
    elif action == "ACKNOWLEDGE_REVIEW":
        source_ws.cell(row, source_headers["manual_update_date"], now.date())
        source_ws.cell(row, source_headers["manual_updated_by"], operator)
    elif action == "APPLY_FIELD_UPDATES":
        updates = source_updates_from_operation(record, payload)
        if not updates:
            raise HumanOperationError("At least one human-governed source field must be completed before this item can be accepted.")
        for field, value in updates.items():
            source_ws.cell(row, source_headers[field], value)
    else:
        raise HumanOperationError(f"Unsupported source action: {action}")
    source_ws.cell(row, source_headers["manual_update_date"], now.date())
    source_ws.cell(row, source_headers["manual_updated_by"], operator)
    source_ws.cell(row, source_headers["notes"], append_note(source_ws.cell(row, source_headers["notes"]).value, f"Human operation {now.date().isoformat()} by {operator}: {note}"))
    return source_id


def process_decisions(config_path: Path) -> dict[str, Any]:
    config = updater.read_config(config_path)
    now = updater.local_now(config)
    retry_operations: list[tuple[str, str]] = []
    processed: list[str] = []
    waiting: list[str] = []
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        ops = wb[human_sheet_name(config)]
        op_headers = operation_headers(ops)
        source_ws = wb[config["sheet_name"]]
        source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
        changed = False
        final_row = last_operation_row(ops, op_headers)
        for row in range(HUMAN_DATA_START_ROW, final_row + 1):
            record = operation_record(ops, row, op_headers)
            operation_id = str(record.get("operation_id") or "").strip()
            program_status = str(record.get("program_status") or "PENDING").strip().upper()
            raw_decision = str(record.get("decision") or "PENDING").strip().upper()
            decision = ACTION_NORMALISATION.get(raw_decision, raw_decision)
            if not operation_id:
                continue
            ops.row_dimensions[row].hidden = program_status in FINAL_STATUSES
            if program_status in FINAL_STATUSES or decision == "PENDING":
                continue
            if decision not in DECISIONS:
                set_operation_fields(ops, row, op_headers, {"program_status": "WAITING_FOR_HUMAN", "program_note": f"Invalid decision: {decision}", "decision": "PENDING"})
                waiting.append(operation_id)
                changed = True
                continue
            operator = str(record.get("operator") or "").strip()
            note = str(record.get("operator_note") or "").strip()
            if not operator or (decision in {"REJECT", "INCORRECT"} and not note):
                set_operation_fields(ops, row, op_headers, {"program_status": "WAITING_FOR_HUMAN", "program_note": "Operator is required; REJECT and INCORRECT also require an operator note.", "decision": "PENDING"})
                waiting.append(operation_id)
                changed = True
                continue
            checked_at = record.get("checked_at") or now
            operation_type = str(record.get("operation_type") or "").strip().upper()
            payload = parse_payload(record.get("payload_json"))
            try:
                if operation_type in {"SOURCE_REVIEW", "SELECTION_REVIEW"}:
                    source_id = str(record.get("source_id") or payload.get("source_id") or "").strip()
                    source_row = find_source_row(source_ws, source_headers, source_id)
                    if source_row is None:
                        raise HumanOperationError(f"Source not found: {source_id}")
                    current_source = updater.record_from_row(source_ws, source_row, source_headers)
                    expected_fingerprint = str(payload.get("source_fingerprint") or "").strip()
                    if expected_fingerprint and source_record_fingerprint(current_source) != expected_fingerprint:
                        raise HumanOperationError("The source changed after this task was created. Run Routine Cycle to refresh the task before applying it.")
                if operation_type == "NEW_SOURCE_CANDIDATE":
                    payload = candidate_payload_from_operation(record, payload)
                    if decision == "ACCEPT":
                        new_id = add_candidate_to_source_register(source_ws, source_headers, payload, config, now, operator, note)
                        result_note = f"Candidate accepted and added to Source Register as {new_id}."
                        final_status = "APPLIED"
                    elif decision == "REJECT":
                        archive_intake_file(payload, config, operation_id, "rejected")
                        result_note = "Candidate rejected; no Source Register row was created."
                        final_status = "REJECTED"
                    else:
                        raise HumanOperationError("New-source candidates require ACCEPT or REJECT.")
                elif operation_type == "MANUAL_FILE_REPLACEMENT":
                    payload = candidate_payload_from_operation(record, payload)
                    if decision == "ACCEPT":
                        source_id = replace_source_with_manual_file(source_ws, source_headers, record, payload, config, now)
                        result_note = f"Authorised manual file accepted as the current snapshot for {source_id}."
                        final_status = "APPLIED"
                    elif decision == "REJECT":
                        archive_intake_file(payload, config, str(record.get("source_id") or "unmatched"), "rejected")
                        result_note = "Manual replacement rejected; the current registered file was preserved."
                        final_status = "REJECTED"
                    else:
                        raise HumanOperationError("Manual replacements require ACCEPT or REJECT.")
                elif operation_type == "RANDOM_QA_CHECK":
                    if decision not in {"CORRECT", "INCORRECT"}:
                        raise HumanOperationError("Random QA items require CORRECT or INCORRECT.")
                    result_note = f"Random QA recorded as {decision}."
                    final_status = "APPLIED"
                    if decision == "INCORRECT":
                        followup_id = "REV-QA-" + hashlib.sha256(f"{record.get('source_id')}|{operation_id}".encode("utf-8")).hexdigest()[:12]
                        followup = {
                            "operation_id": followup_id,
                            "operation_type": "SOURCE_REVIEW",
                            "source_id": record.get("source_id"),
                            "source_title": record.get("source_title"),
                            "trigger": "RANDOM_QA_FAILURE",
                            "proposed_action": "APPLY_FIELD_UPDATES",
                            "decision": "PENDING",
                            "program_status": "PENDING",
                            "program_note": f"Random QA found an issue: {note}. The review agent must propose structured field updates.",
                            "official_url": record.get("official_url"),
                            "issuer": record.get("issuer"),
                            "jurisdiction": record.get("jurisdiction"),
                            "document_type": record.get("document_type"),
                            "source_family": record.get("source_family"),
                            "requirement_role": record.get("requirement_role"),
                            "authoritative_language": record.get("authoritative_language"),
                            "file_format": record.get("file_format"),
                            "created_at": now,
                            "payload_json": json.dumps({"action": "APPLY_FIELD_UPDATES", "source_id": record.get("source_id"), "updates": {}}, sort_keys=True),
                        }
                        for operation_field in OPERATION_SOURCE_FIELD_MAP:
                            followup[operation_field] = record.get(operation_field)
                        append_operation(ops, op_headers, followup)
                elif operation_type in {"SOURCE_REVIEW", "SELECTION_REVIEW"}:
                    if decision == "REJECT":
                        set_operation_fields(ops, row, op_headers, {"decision": "PENDING", "checked_at": checked_at, "program_status": "NEEDS_REPLAN", "program_operated_at": now, "program_note": f"Proposed action rejected by {operator}: {note}"})
                        waiting.append(operation_id)
                        changed = True
                        continue
                    if decision != "ACCEPT":
                        raise HumanOperationError("Source and selection review items require ACCEPT or REJECT.")
                    selection_review_required = operation_type == "SELECTION_REVIEW" or "SELECTION_PENDING" in set(payload.get("issue_codes") or [])
                    if selection_review_required:
                        selection_decision = str(record.get("operator_selection_decision") or "PENDING").strip().upper()
                        if selection_decision not in {"INCLUDE", "PENDING", "EXCLUDE"}:
                            raise HumanOperationError("Selection review requires operator_selection_decision INCLUDE, PENDING or EXCLUDE.")
                        if selection_decision == "PENDING" and not note:
                            raise HumanOperationError("A selection kept as PENDING requires an operator note explaining what remains unresolved.")
                    action = str(payload.get("action") or record.get("proposed_action") or "").strip().upper()
                    if action == "RETRY_SOURCE":
                        source_id = str(record.get("source_id") or "").strip()
                        if not source_id:
                            raise HumanOperationError("RETRY_SOURCE requires source_id.")
                        retry_operations.append((operation_id, source_id))
                        set_operation_fields(ops, row, op_headers, {"checked_at": checked_at, "program_status": "QUEUED", "program_operated_at": now, "program_note": "Accepted; automated retry queued."})
                        changed = True
                        continue
                    apply_source_action(source_ws, source_headers, record, payload, now)
                    if bool(payload.get("retry_after_apply")):
                        source_id = str(record.get("source_id") or "").strip()
                        retry_operations.append((operation_id, source_id))
                        set_operation_fields(ops, row, op_headers, {"checked_at": checked_at, "program_status": "QUEUED", "program_operated_at": now, "program_note": "Updates applied; an automated retry is queued."})
                        changed = True
                        continue
                    result_note = f"Accepted action {action} was applied to Source Register."
                    final_status = "APPLIED"
                else:
                    raise HumanOperationError(f"Unsupported operation_type: {operation_type}")
                set_operation_fields(ops, row, op_headers, {"checked_at": checked_at, "program_status": final_status, "program_operated_at": now, "program_note": result_note})
                ops.row_dimensions[row].hidden = True
                processed.append(operation_id)
                changed = True
            except HumanOperationError as exc:
                set_operation_fields(ops, row, op_headers, {"decision": "PENDING", "checked_at": checked_at, "program_status": "WAITING_FOR_HUMAN", "program_operated_at": now, "program_note": str(exc)})
                ops.row_dimensions[row].hidden = False
                waiting.append(operation_id)
                changed = True
        update_operation_table(ops, last_operation_row(ops, op_headers))
        if changed:
            prepare_human_workbook(wb, ops, op_headers)
            updater.create_backup(config["workbook"], config["backup_root"], now)
            prepare_human_workbook(wb, ops, op_headers)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()

    retry_results: dict[str, str] = {}
    for operation_id, source_id in retry_operations:
        updater.run_updates(config, {source_id}, include_pending=True, available_only=False)
        with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
            expected_mtime = updater.workbook_mtime(config["workbook"])
            wb = load_workbook(config["workbook"], data_only=False)
            ops = wb[human_sheet_name(config)]
            op_headers = operation_headers(ops)
            row = operation_index(ops, op_headers)[operation_id]
            source_ws = wb[config["sheet_name"]]
            source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
            source_row = find_source_row(source_ws, source_headers, source_id)
            status = str(source_ws.cell(source_row, source_headers["download_status"]).value or "").strip().upper() if source_row else "FAIL"
            if status == "SUCCESS":
                set_operation_fields(ops, row, op_headers, {"program_status": "APPLIED", "program_operated_at": updater.local_now(config), "program_note": "Accepted retry completed successfully."})
                ops.row_dimensions[row].hidden = True
                processed.append(operation_id)
                retry_results[source_id] = "SUCCESS"
            else:
                reason = source_ws.cell(source_row, source_headers["failure_reason"]).value if source_row else "Source row missing after retry."
                set_operation_fields(ops, row, op_headers, {"decision": "PENDING", "program_status": "WAITING_FOR_HUMAN", "program_operated_at": updater.local_now(config), "program_note": f"Retry failed: {reason}"})
                ops.row_dimensions[row].hidden = False
                waiting.append(operation_id)
                retry_results[source_id] = "FAIL"
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
            wb.close()
    return {
        "processed_count": len(set(processed)),
        "waiting_count": len(set(waiting)),
        "retry_results": retry_results,
    }


def archive_completed_rows(config_path: Path) -> dict[str, Any]:
    config = updater.read_config(config_path)
    changed = 0
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        ws = wb[human_sheet_name(config)]
        headers = operation_headers(ws)
        for row in range(HUMAN_DATA_START_ROW, last_operation_row(ws, headers) + 1):
            hidden = str(ws.cell(row, headers["program_status"]).value or "").strip().upper() in FINAL_STATUSES
            if bool(ws.row_dimensions[row].hidden) != hidden:
                ws.row_dimensions[row].hidden = hidden
                changed += 1
        if changed:
            prepare_human_workbook(wb, ws, headers)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()
    return {"row_visibility_updates": changed}


def review_trigger_is_resolved(trigger: str, source: dict[str, Any]) -> bool:
    trigger = trigger.strip().upper()
    if ";" in trigger:
        codes = [code.strip() for code in trigger.split(";") if code.strip()]
        return bool(codes) and all(review_trigger_is_resolved(code, source) for code in codes)
    if trigger == "DOWNLOAD_FAILURE":
        return (
            str(source.get("download_status") or "").strip().upper() == "SUCCESS"
            and str(source.get("snapshot_status") or "").strip().upper() == "STORED"
        )
    if trigger == "MISSING_FILE":
        return str(source.get("snapshot_status") or "").strip().upper() == "STORED"
    if trigger == "PROVENANCE_UNVERIFIED":
        return str(source.get("provenance_status") or "").strip().upper() not in {"", "UNVERIFIED"}
    if trigger == "APPLICABILITY_UNCONFIRMED":
        return bool(str(source.get("applicability_reference") or "").strip())
    return False


def reconcile_resolved_reviews(config_path: Path) -> dict[str, Any]:
    """Archive review rows whose underlying program state is already resolved."""
    config = updater.read_config(config_path)
    now = updater.local_now(config)
    resolved: list[str] = []
    with updater.updater_lock(config["workbook"], float(config["lock_stale_hours"])):
        expected_mtime = updater.workbook_mtime(config["workbook"])
        wb = load_workbook(config["workbook"], data_only=False)
        ops = wb[human_sheet_name(config)]
        op_headers = operation_headers(ops)
        source_ws = wb[config["sheet_name"]]
        source_headers = updater.workbook_headers(source_ws, int(config["header_row"]))
        sources: dict[str, dict[str, Any]] = {}
        for source_row in range(int(config["data_start_row"]), source_ws.max_row + 1):
            source = updater.record_from_row(source_ws, source_row, source_headers)
            source_id = str(source.get("source_id") or "").strip()
            if source_id:
                sources[source_id] = source
        for row in range(HUMAN_DATA_START_ROW, last_operation_row(ops, op_headers) + 1):
            record = operation_record(ops, row, op_headers)
            if str(record.get("operation_type") or "").strip().upper() != "SOURCE_REVIEW":
                continue
            if str(record.get("program_status") or "PENDING").strip().upper() not in OPEN_STATUSES:
                continue
            if str(record.get("decision") or "PENDING").strip().upper() != "PENDING":
                continue
            source_id = str(record.get("source_id") or "").strip()
            source = sources.get(source_id)
            if not source or not review_trigger_is_resolved(str(record.get("trigger") or ""), source):
                continue
            operation_id = str(record.get("operation_id") or "").strip()
            set_operation_fields(ops, row, op_headers, {
                "program_status": "APPLIED",
                "program_operated_at": now,
                "program_note": "The underlying source state was resolved automatically; no human decision was required.",
            })
            ops.row_dimensions[row].hidden = True
            resolved.append(operation_id)
        if resolved:
            prepare_human_workbook(wb, ops, op_headers)
            updater.create_backup(config["workbook"], config["backup_root"], now)
            updater.save_workbook_atomic(wb, config["workbook"], expected_mtime)
        wb.close()
    return {"resolved_count": len(resolved), "operation_ids": resolved}


def run_human_cycle(config_path: Path) -> dict[str, Any]:
    intake = manual_intake.scan_manual_intake(config_path)
    return {
        "manual_intake": intake,
        "decision_processing": process_decisions(config_path),
        "resolved_review_reconciliation": reconcile_resolved_reviews(config_path),
        "discovery_gate": sync_discovery_candidates(config_path),
        "archive": archive_completed_rows(config_path),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Process Human Operation Desktop decisions and requests.")
    parser.add_argument("--config", default="config/config.json")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = run_human_cycle(Path(args.config).expanduser().resolve())
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
