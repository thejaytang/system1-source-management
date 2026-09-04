#!/usr/bin/env python3
"""Turn files dropped by humans into reviewable source candidates.

The intake agent is deliberately conservative.  It extracts local metadata and
embedded URLs without transmitting private file content to a public search
service.  When an equivalent official HTML source is identifiable, HTML is
ranked before PDF.  Missing fields remain visible for human completion in the
Human Operation Desktop.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from openpyxl import load_workbook

import source_updater as updater


SUPPORTED_EXTENSIONS = {".html", ".htm", ".pdf", ".xlsx", ".zip"}
OFFICIAL_DOMAIN_HINTS = (
    "lovdata.no", "regjeringen.no", "fiskeridir.no", "mattilsynet.no",
    "miljodirektoratet.no", "arbeidstilsynet.no", "eur-lex.europa.eu",
    "standard.no", "iso.org", "asc-aqua.org", "globalgap.org",
)
ISSUER_BY_DOMAIN = {
    "lovdata.no": "Lovdata",
    "regjeringen.no": "Norwegian Government",
    "fiskeridir.no": "Norwegian Directorate of Fisheries",
    "mattilsynet.no": "Norwegian Food Safety Authority",
    "miljodirektoratet.no": "Norwegian Environment Agency",
    "arbeidstilsynet.no": "Norwegian Labour Inspection Authority",
    "eur-lex.europa.eu": "European Union",
    "standard.no": "Standards Norway",
    "iso.org": "International Organization for Standardization",
    "asc-aqua.org": "Aquaculture Stewardship Council",
    "globalgap.org": "GLOBALG.A.P.",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def filename_title(path: Path) -> str:
    title = re.sub(r"[_-]+", " ", path.stem)
    title = re.sub(r"\s+", " ", title).strip()
    title = re.sub(r"^[A-Z]{2}\d{3}(?:-\d{3})?\s*", "", title)
    return title or path.stem


def local_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".html", ".htm"}:
        return path.read_text(encoding="utf-8", errors="replace")[:1_000_000]
    if suffix == ".pdf":
        return path.read_bytes()[:2_000_000].decode("latin-1", errors="ignore")
    if suffix in {".xlsx", ".zip"} and zipfile.is_zipfile(path):
        parts: list[str] = []
        with zipfile.ZipFile(path) as archive:
            for member in ("docProps/core.xml", "xl/workbook.xml", "[Content_Types].xml"):
                if member in archive.namelist():
                    parts.append(archive.read(member)[:200_000].decode("utf-8", errors="replace"))
        return "\n".join(parts)
    return ""


def extracted_title(path: Path, text: str) -> str:
    if path.suffix.lower() in {".html", ".htm"}:
        match = re.search(r"<title[^>]*>(.*?)</title>", text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            cleaned = re.sub(r"<[^>]+>|\s+", " ", html.unescape(match.group(1))).strip()
            if cleaned:
                return cleaned[:250]
    if path.suffix.lower() == ".pdf":
        match = re.search(r"/Title\s*\((.{3,250}?)\)", text, flags=re.DOTALL)
        if match:
            cleaned = re.sub(r"\\[()\\]", "", match.group(1)).strip()
            if cleaned:
                return cleaned[:250]
    return filename_title(path)


def extract_urls(text: str) -> list[str]:
    raw = re.findall(r"https?://[^\s<>\"'()\[\]{}]+", html.unescape(text), flags=re.IGNORECASE)
    result: list[str] = []
    for value in raw:
        cleaned = value.rstrip(".,;:!?")
        if cleaned not in result:
            result.append(cleaned)
    return result[:100]


def url_format(url: str) -> str:
    suffix = Path(urlparse(url).path).suffix.lower()
    return {".pdf": "pdf", ".xlsx": "xlsx", ".zip": "zip"}.get(suffix, "html")


def normalise_format(value: str) -> str:
    """Normalise supported extensions without turning ``html`` into ``htmll``."""
    value = str(value or "").strip().lower().lstrip(".")
    return "html" if value in {"htm", "html"} else value


def official_domain(url: str) -> str:
    host = (urlparse(url).hostname or "").lower()
    for domain in OFFICIAL_DOMAIN_HINTS:
        if host == domain or host.endswith("." + domain):
            return domain
    return ""


def choose_official_url(urls: list[str], format_order: list[str]) -> str:
    candidates = [url for url in urls if official_domain(url)]
    rank = {value: index for index, value in enumerate(format_order)}
    candidates.sort(key=lambda url: (rank.get(url_format(url), 99), len(url)))
    return candidates[0] if candidates else ""


def infer_language(text: str) -> str:
    sample = re.sub(r"[^A-Za-zÆØÅæøå ]", " ", text[:200_000]).lower()
    norwegian = sum(sample.count(token) for token in (" og ", " forskrift", " krav", " gjelder", " skal ", " til "))
    english = sum(sample.count(token) for token in (" and ", " regulation", " requirements", " shall ", " applies ", " standard"))
    if norwegian > english * 1.2:
        return "Norwegian"
    if english > norwegian * 1.2:
        return "English"
    return "Other"


def infer_classification(title: str, url: str) -> tuple[str, str]:
    domain = official_domain(url)
    lowered = title.lower()
    if domain in {"standard.no", "iso.org"}:
        return "Standards body", "B_Standards_Body"
    if domain in {"asc-aqua.org", "globalgap.org"}:
        return "Certification scheme", "C_Certification_Scheme"
    if domain:
        return "Public authority", "A_Public_Authority"
    if any(token in lowered for token in ("manual", "installation", "operation manual", "datasheet")):
        return "Manufacturer / supplier", "D_Manufacturer_Supplier"
    return "Pending classification", "Z_Pending_Classification"


def infer_document_type(title: str) -> tuple[str, str]:
    lowered = title.lower()
    if "regulation" in lowered or "forskrift" in lowered:
        return "Regulation", "Primary normative"
    if re.search(r"\bact\b|\blov\b", lowered):
        return "Law", "Primary normative"
    if "standard" in lowered:
        return "Technical standard", "Referenced normative"
    if "manual" in lowered:
        return "Equipment manual / OEM instruction", "Asset-specific normative"
    if "audit" in lowered or "checklist" in lowered:
        return "Audit / assessment document", "Audit / assessment tool"
    if "guid" in lowered or "veileder" in lowered:
        return "Interpretation document", "Guidance / interpretation"
    return "Other", "Draft working material"


def load_inbox(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [item for item in data if isinstance(item, dict)] if isinstance(data, list) else []


def existing_hashes(config: dict[str, Any], candidates: list[dict[str, Any]]) -> set[str]:
    hashes = {str(item.get("content_hash") or "").strip().lower() for item in candidates}
    # Normal mode is required because some valid workbooks expose no max_row
    # through openpyxl's streaming reader.
    workbook = load_workbook(config["workbook"], data_only=False)
    sheet = workbook[config["sheet_name"]]
    headers = updater.workbook_headers(sheet, int(config["header_row"]))
    for row in range(int(config["data_start_row"]), sheet.max_row + 1):
        value = str(sheet.cell(row, headers["content_hash"]).value or "").strip().lower()
        if value:
            hashes.add(value)
    workbook.close()
    return hashes


def candidate_from_file(path: Path, config: dict[str, Any]) -> dict[str, Any]:
    text = local_text(path)
    title = extracted_title(path, text)
    urls = extract_urls(text)
    format_order = [str(value).lower() for value in config.get("format_preference", ["html", "pdf", "xlsx", "zip"])]
    preferred_url = choose_official_url(urls, format_order)
    source_family, folder_code = infer_classification(title, preferred_url)
    document_type, requirement_role = infer_document_type(title)
    domain = official_domain(preferred_url)
    chosen_format = url_format(preferred_url) if preferred_url else normalise_format(path.suffix)
    paywall = domain in {"standard.no", "iso.org"}
    return {
        "candidate_origin": "HUMAN_DROP",
        "intake_file": path.name,
        "content_hash": sha256_file(path),
        "source_title": title,
        "official_url": preferred_url,
        "retrieval_url": preferred_url,
        "embedded_url_candidates": urls,
        "official_search_query": title,
        "folder_code": folder_code,
        "file_format": chosen_format,
        "manual_file_format": normalise_format(path.suffix),
        "issuer": ISSUER_BY_DOMAIN.get(domain),
        "jurisdiction": "Norway" if domain.endswith(".no") else ("EU / EEA" if domain == "eur-lex.europa.eu" else "International"),
        "authoritative_language": infer_language(text),
        "source_family": source_family,
        "document_type": document_type,
        "requirement_role": requirement_role,
        "inclusion_rationale": "Human-supplied candidate; confirm official-source equivalence and scope before inclusion.",
        "authority_quality": "MEDIUM",
        "scope_relevance": "MEDIUM",
        "version_currency": "MEDIUM",
        "traceability": "MEDIUM",
        "access_permission": "MEDIUM" if paywall else "HIGH",
        "automation_readiness": "HIGH" if preferred_url and not paywall else "LOW",
        "operator_selection_decision": "PENDING",
        "acquisition_channel": "OTHER",
        "provenance_status": "UNVERIFIED",
        "update_model": "UNKNOWN",
        "notes": "Created by the manual-intake agent from a file placed in Data/00_Human_Intake.",
    }


def scan_manual_intake(config_path: Path) -> dict[str, Any]:
    config = updater.read_config(config_path)
    root = config.get("manual_intake_root")
    if root is None:
        return {"scanned_count": 0, "added_count": 0, "candidates": []}
    root.mkdir(parents=True, exist_ok=True)
    inbox_path = (config_path.parent / str(config.get("discovery_candidate_inbox") or "../runtime/inbox/discovery_candidates.json")).resolve()
    candidates = load_inbox(inbox_path)
    hashes = existing_hashes(config, candidates)
    added: list[dict[str, Any]] = []
    for path in sorted(root.iterdir()):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS or path.name.startswith("."):
            continue
        digest = sha256_file(path)
        if digest in hashes:
            continue
        candidate = candidate_from_file(path, config)
        candidates.append(candidate)
        added.append(candidate)
        hashes.add(digest)
    if added:
        inbox_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=inbox_path.parent, prefix=".discovery-", suffix=".json") as handle:
            json.dump(candidates, handle, ensure_ascii=False, indent=2)
            temp_path = Path(handle.name)
        temp_path.replace(inbox_path)
    return {"scanned_count": len(candidates), "added_count": len(added), "candidates": [item["intake_file"] for item in added]}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Scan Data/00_Human_Intake and create human-review candidates.")
    parser.add_argument("--config", default="config/config.json")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = scan_manual_intake(Path(args.config).expanduser().resolve())
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
