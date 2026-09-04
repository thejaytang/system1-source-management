#!/usr/bin/env python3
"""Requirement Source Registry automated updater.

The workbook is the human-facing source of truth. This program reads the stable
field names from row 2 of ``Source Register`` and only updates program-owned fields.
New downloads are validated in a temporary directory before any current file is
changed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from system1.workbook_guard import WorkbookBusyError, system_lock


MANDATORY_HEADERS = {
    "source_id",
    "snapshot_id",
    "source_title",
    "retrieval_url",
    "folder_code",
    "file_format",
    "stored_filename",
    "snapshot_status",
    "issuer",
    "source_family",
    "requirement_role",
    "authority_quality",
    "scope_relevance",
    "version_currency",
    "traceability",
    "access_permission",
    "automation_readiness",
    "operator_selection_decision",
    "acquisition_channel",
    "provenance_status",
    "update_model",
    "download_status",
    "last_attempt_at",
    "last_success_at",
    "content_change",
    "current_origin",
    "failure_stage",
    "failure_reason",
    "content_hash",
    "source_status",
}
SCORE_FIELDS = (
    "authority_quality",
    "scope_relevance",
    "version_currency",
    "traceability",
    "access_permission",
)
PROGRAM_FIELDS = {
    "snapshot_id",
    "stored_filename",
    "snapshot_status",
    "download_status",
    "last_attempt_at",
    "last_success_at",
    "content_change",
    "current_origin",
    "failure_stage",
    "failure_reason",
    "content_hash",
}
SOURCE_ID_RE = re.compile(r"^[A-Z]{2}\d{3}$")


class UpdaterError(RuntimeError):
    def __init__(self, stage: str, message: str):
        super().__init__(message)
        self.stage = stage


@dataclass
class DownloadResult:
    path: Path
    sha256: str
    byte_count: int
    mime_type: str
    final_url: str


def read_config(config_path: Path) -> dict[str, Any]:
    with config_path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    base = config_path.parent.resolve()
    for key in ("workbook", "source_root", "backup_root", "log_root", "manual_intake_root", "manual_intake_archive_root"):
        if key not in config:
            continue
        config[key] = (base / config[key]).resolve()
    config.setdefault("sheet_name", "Source Register")
    config.setdefault("header_row", 2)
    config.setdefault("data_start_row", 3)
    config.setdefault("timezone", "Europe/Oslo")
    config.setdefault("timeout_seconds", 30)
    config.setdefault("max_bytes", 50 * 1024 * 1024)
    config.setdefault("minimum_pdf_bytes", 512)
    config.setdefault("minimum_html_bytes", 100)
    config.setdefault("minimum_binary_bytes", 100)
    config.setdefault("user_agent", "RequirementSourceUpdater/1.0")
    config.setdefault("format_preference", ["html", "pdf", "xlsx", "zip"])
    config.setdefault("language_preference", {
        "mode": "authoritative_first",
        "working_language_order": ["English", "Norwegian", "Other"],
    })
    config.setdefault("eligible_source_statuses", ["CURRENT"])
    config.setdefault("lock_stale_hours", 12)
    config.setdefault("folder_prefix_map", {
        "A_Public_Authority": "PA",
        "B_Standards_Body": "SB",
        "C_Certification_Scheme": "CS",
        "D_Manufacturer_Supplier": "MS",
        "E_Project_Engineering": "PE",
        "F_Client_Internal": "CI",
        "Z_Pending_Classification": "PC",
    })
    return config


def local_now(config: dict[str, Any]) -> datetime:
    return datetime.now(ZoneInfo(config["timezone"])).replace(tzinfo=None, microsecond=0)


def iso_timestamp(config: dict[str, Any]) -> str:
    return datetime.now(ZoneInfo(config["timezone"])).isoformat(timespec="seconds")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def selection_from_scores(record: dict[str, Any]) -> str:
    """Return the effective source-selection state.

    The operator owns the intent to INCLUDE/PENDING/EXCLUDE.  The program owns
    the effective status and will not return INCLUDE until a validated current
    file exists and the acquisition status is SUCCESS.
    """
    operator_decision = str(record.get("operator_selection_decision") or "PENDING").strip().upper()
    scores = [str(record.get(name) or "").strip().upper() for name in SCORE_FIELDS]
    if operator_decision == "EXCLUDE" or "LOW" in scores:
        return "EXCLUDE"
    issuer = str(record.get("issuer") or "").strip()
    acquisition = str(record.get("acquisition_channel") or "").strip().upper()
    provenance = str(record.get("provenance_status") or "").strip().upper()
    if not issuer or not acquisition or not provenance or provenance == "UNVERIFIED":
        return "PENDING"
    if acquisition == "OFFICIAL_WEBSITE" and not str(record.get("official_url") or "").strip():
        return "PENDING"
    document_type = str(record.get("document_type") or "").strip()
    requirement_role = str(record.get("requirement_role") or "").strip()
    if document_type == "Equipment manual / OEM instruction" or requirement_role == "Asset-specific normative":
        if (
            str(record.get("source_family") or "").strip() != "Manufacturer / supplier"
            or not str(record.get("applicability_reference") or "").strip()
            or not str(record.get("inclusion_rationale") or "").strip()
        ):
            return "PENDING"
    if operator_decision != "INCLUDE":
        return "PENDING"
    if str(record.get("download_status") or "").strip().upper() != "SUCCESS":
        return "PENDING"
    if str(record.get("snapshot_status") or "").strip().upper() != "STORED":
        return "PENDING"
    if scores == ["HIGH"] * len(SCORE_FIELDS):
        return "INCLUDE"
    return "PENDING"


def workbook_headers(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        value = str(cell.value or "").strip()
        if value:
            if value in headers:
                raise UpdaterError("CONFIG", f"Duplicate field name: {value}")
            headers[value] = cell.column
    missing = sorted(MANDATORY_HEADERS - headers.keys())
    if missing:
        raise UpdaterError("CONFIG", f"Source Register is missing fields: {', '.join(missing)}")
    return headers


def record_from_row(ws, row: int, headers: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row=row, column=column).value for name, column in headers.items()}


def set_fields(ws, row: int, headers: dict[str, int], updates: dict[str, Any]) -> None:
    unexpected = set(updates) - PROGRAM_FIELDS
    if unexpected:
        raise UpdaterError("CONFIG", f"The updater attempted to modify non-program fields: {sorted(unexpected)}")
    for name, value in updates.items():
        ws.cell(row=row, column=headers[name], value=value)


def workbook_mtime(path: Path) -> int:
    return path.stat().st_mtime_ns


def save_workbook_atomic(wb, workbook_path: Path, expected_mtime_ns: int) -> int:
    if workbook_mtime(workbook_path) != expected_mtime_ns:
        raise UpdaterError("STORAGE", "The workbook changed while the updater was running; the write was stopped to protect human edits.")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}.", suffix=".tmp.xlsx", dir=workbook_path.parent
    )
    os.close(fd)
    temporary_path = Path(temporary_name)
    try:
        if getattr(wb, "calculation", None) is not None:
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        wb.save(temporary_path)
        with zipfile.ZipFile(temporary_path) as archive:
            bad_member = archive.testzip()
            if bad_member:
                raise UpdaterError("STORAGE", f"Temporary Excel file is corrupt: {bad_member}")
        try:
            os.replace(temporary_path, workbook_path)
        except PermissionError as exc:
            raise UpdaterError("STORAGE", "Excel or another program is using the workbook; the validated temporary copy was not promoted.") from exc
    finally:
        temporary_path.unlink(missing_ok=True)
    return workbook_mtime(workbook_path)


def create_backup(workbook_path: Path, backup_root: Path, now: datetime) -> Path:
    backup_root.mkdir(parents=True, exist_ok=True)
    target = backup_root / f"{workbook_path.stem}_{now:%Y%m%d_%H%M%S}.xlsx"
    counter = 1
    while target.exists():
        target = backup_root / f"{workbook_path.stem}_{now:%Y%m%d_%H%M%S}_{counter}.xlsx"
        counter += 1
    shutil.copy2(workbook_path, target)
    return target


def process_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    return True


@contextmanager
def updater_lock(workbook_path: Path, stale_hours: float):
    try:
        with system_lock(workbook_path, stale_hours):
            yield
    except WorkbookBusyError as exc:
        raise UpdaterError("STORAGE", str(exc)) from exc


def append_log(config: dict[str, Any], event: dict[str, Any]) -> None:
    log_root: Path = config["log_root"]
    log_root.mkdir(parents=True, exist_ok=True)
    event = {"timestamp": iso_timestamp(config), **event}
    log_path = log_root / f"updater_{datetime.now(ZoneInfo(config['timezone'])):%Y-%m-%d}.jsonl"
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False, default=str) + "\n")


def check_url(url: str) -> None:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise UpdaterError("CONFIG", "retrieval_url must be a complete http or https URL.")


def read_limited(response, destination: Path, max_bytes: int) -> tuple[str, int]:
    content_length = response.headers.get("Content-Length")
    if content_length:
        try:
            if int(content_length) > max_bytes:
                raise UpdaterError("DOWNLOAD", f"File exceeds the configured limit of {max_bytes} bytes.")
        except ValueError:
            pass
    digest = hashlib.sha256()
    total = 0
    try:
        with destination.open("wb") as handle:
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise UpdaterError("DOWNLOAD", f"File exceeds the configured limit of {max_bytes} bytes.")
                digest.update(chunk)
                handle.write(chunk)
    except UpdaterError:
        raise
    except (OSError, TimeoutError) as exc:
        raise UpdaterError("DOWNLOAD", f"Download did not complete: {exc}") from exc
    return digest.hexdigest(), total


def validate_pdf(path: Path, byte_count: int, minimum_bytes: int) -> None:
    if byte_count < minimum_bytes:
        raise UpdaterError("CONTENT", f"PDF content is too small: {byte_count} bytes.")
    with path.open("rb") as handle:
        start = handle.read(1024)
        handle.seek(max(0, byte_count - 4096))
        end = handle.read()
    if b"%PDF-" not in start:
        raise UpdaterError("FORMAT", "The file has no PDF signature; the response may be an HTML or error page.")
    if b"%%EOF" not in end:
        raise UpdaterError("DOWNLOAD", "The PDF has no EOF marker and may be incomplete.")


def validate_html(path: Path, byte_count: int, minimum_bytes: int) -> None:
    if byte_count < minimum_bytes:
        raise UpdaterError("CONTENT", f"HTML content is too small: {byte_count} bytes.")
    sample = path.read_bytes()[:256_000]
    if sample.lstrip().startswith(b"%PDF-"):
        raise UpdaterError("FORMAT", "Expected HTML, but the downloaded content is a PDF.")
    text = sample.decode("utf-8", errors="replace")
    lowered = text.lower()
    if not any(marker in lowered for marker in ("<html", "<!doctype html", "<body")):
        raise UpdaterError("FORMAT", "The content has no recognisable HTML structure.")
    title_match = re.search(r"<title[^>]*>(.*?)</title>", lowered, flags=re.DOTALL)
    title = re.sub(r"\s+", " ", title_match.group(1)).strip() if title_match else ""
    blocked_titles = ("login", "sign in", "access denied", "forbidden", "not found", "error")
    if title and any(marker in title for marker in blocked_titles):
        raise UpdaterError("CONTENT", f"The page title indicates that this is not the target content: {title[:120]}")
    visible = re.sub(r"<script\b[^>]*>.*?</script>|<style\b[^>]*>.*?</style>|<[^>]+>", " ", lowered, flags=re.DOTALL)
    visible = re.sub(r"\s+", " ", visible).strip()
    if len(visible) < minimum_bytes // 2:
        raise UpdaterError("CONTENT", "HTML readable body text is too short; the response may be a redirect or error page.")


def validate_zip_container(path: Path, byte_count: int, minimum_bytes: int, require_xlsx: bool = False) -> None:
    label = "XLSX" if require_xlsx else "ZIP"
    if byte_count < minimum_bytes:
        raise UpdaterError("CONTENT", f"{label} content is too small: {byte_count} bytes.")
    if not zipfile.is_zipfile(path):
        raise UpdaterError("FORMAT", f"The file has no valid {label} / ZIP signature.")
    try:
        with zipfile.ZipFile(path) as archive:
            bad_member = archive.testzip()
            if bad_member:
                raise UpdaterError("DOWNLOAD", f"Archive member is corrupt: {bad_member}")
            if require_xlsx:
                names = set(archive.namelist())
                required = {"[Content_Types].xml", "xl/workbook.xml"}
                missing = required - names
                if missing:
                    raise UpdaterError("FORMAT", f"XLSX is missing required structures: {', '.join(sorted(missing))}")
    except zipfile.BadZipFile as exc:
        raise UpdaterError("FORMAT", f"Unable to read the {label} archive structure.") from exc


def request_accept_header(file_format: str) -> str:
    """Prefer the registered original format during HTTP content negotiation."""
    return {
        "pdf": "application/pdf,*/*;q=0.1",
        "html": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.1",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/zip;q=0.9,*/*;q=0.1",
        "zip": "application/zip,application/octet-stream;q=0.9,*/*;q=0.1",
    }[file_format]


def download_and_validate(record: dict[str, Any], config: dict[str, Any], temp_dir: Path) -> DownloadResult:
    url = str(record.get("retrieval_url") or "").strip()
    file_format = str(record.get("file_format") or "").strip().lower()
    if not url:
        raise UpdaterError("CONFIG", "retrieval_url is missing.")
    if file_format not in {"pdf", "html", "xlsx", "zip"}:
        raise UpdaterError("CONFIG", "file_format must be pdf, html, xlsx or zip.")
    check_url(url)
    request = urllib.request.Request(url, headers={
        "User-Agent": config["user_agent"],
        "Accept": request_accept_header(file_format),
    })
    destination = temp_dir / f"download.{file_format}"
    try:
        response = urllib.request.urlopen(request, timeout=float(config["timeout_seconds"]))
    except urllib.error.HTTPError as exc:
        try:
            exc.close()
        finally:
            raise UpdaterError("HTTP", f"HTTP {exc.code}: {exc.reason}") from exc
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        raise UpdaterError("CONNECTION", f"Connection failed: {exc}") from exc
    with response:
        status = getattr(response, "status", 200)
        if not (200 <= int(status) < 300):
            raise UpdaterError("HTTP", f"HTTP {status}")
        mime_type = response.headers.get_content_type().lower()
        final_url = response.geturl()
        digest, total = read_limited(response, destination, int(config["max_bytes"]))
    if file_format == "pdf":
        if mime_type.startswith("text/html"):
            raise UpdaterError("FORMAT", "Expected PDF, but the server returned text/html.")
        validate_pdf(destination, total, int(config["minimum_pdf_bytes"]))
    elif file_format == "html":
        if mime_type == "application/pdf":
            raise UpdaterError("FORMAT", "Expected HTML, but the server returned application/pdf.")
        validate_html(destination, total, int(config["minimum_html_bytes"]))
    else:
        if mime_type.startswith("text/html") or mime_type == "application/pdf":
            raise UpdaterError("FORMAT", f"Expected {file_format.upper()}, but the server returned {mime_type}.")
        validate_zip_container(
            destination,
            total,
            int(config["minimum_binary_bytes"]),
            require_xlsx=file_format == "xlsx",
        )
    return DownloadResult(destination, digest, total, mime_type, final_url)


def validate_identity(record: dict[str, Any], folder_prefix_map: dict[str, str] | None = None) -> None:
    source_id = str(record.get("source_id") or "").strip()
    snapshot_id = str(record.get("snapshot_id") or "").strip()
    folder_code = str(record.get("folder_code") or "").strip()
    file_format = str(record.get("file_format") or "").strip().lower()
    stored_filename = str(record.get("stored_filename") or "").strip()
    if not SOURCE_ID_RE.fullmatch(source_id):
        raise UpdaterError("CONFIG", f"Invalid source_id format: {source_id!r}")
    if not re.fullmatch(re.escape(source_id) + r"-\d{3}", snapshot_id):
        raise UpdaterError("CONFIG", f"snapshot_id does not match source_id: {snapshot_id!r}")
    if not folder_code or Path(folder_code).name != folder_code:
        raise UpdaterError("CONFIG", f"folder_code must be a single folder name: {folder_code!r}")
    if folder_prefix_map is not None:
        expected_prefix = folder_prefix_map.get(folder_code)
        if expected_prefix is None:
            raise UpdaterError("CONFIG", f"folder_code is not in the controlled folder mapping: {folder_code!r}")
        if not source_id.startswith(expected_prefix):
            raise UpdaterError(
                "CONFIG",
                f"source_id prefix must be {expected_prefix} to match folder_code {folder_code}.",
            )
    if file_format not in {"pdf", "html", "xlsx", "zip"}:
        raise UpdaterError("CONFIG", f"file_format must be pdf, html, xlsx or zip: {file_format!r}")
    if not stored_filename.startswith(snapshot_id + "_"):
        raise UpdaterError("CONFIG", "stored_filename must start with snapshot_id followed by an underscore.")
    if Path(stored_filename).suffix.lower() != f".{file_format}":
        raise UpdaterError("CONFIG", "stored_filename extension must match file_format.")


def next_snapshot_id(source_id: str, snapshot_id: str) -> str:
    try:
        sequence = int(snapshot_id.rsplit("-", 1)[1]) + 1
    except (IndexError, ValueError) as exc:
        raise UpdaterError("CONFIG", f"Unable to increment snapshot_id: {snapshot_id!r}") from exc
    if sequence > 999:
        raise UpdaterError("CONFIG", f"snapshot_id exceeds the three-digit sequence limit: {snapshot_id!r}")
    return f"{source_id}-{sequence:03d}"


def safe_filename_title(title: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", title.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned[:120] or "source"


def filename_for_snapshot(record: dict[str, Any], snapshot_id: str) -> str:
    old_name = str(record.get("stored_filename") or "").strip()
    extension = str(record["file_format"]).lower()
    old_snapshot = str(record.get("snapshot_id") or "")
    if old_name.startswith(old_snapshot + "_"):
        suffix = old_name[len(old_snapshot) + 1 :]
        return f"{snapshot_id}_{suffix}"
    return f"{snapshot_id}_{safe_filename_title(str(record.get('source_title') or 'source'))}.{extension}"


def archive_current(current_path: Path, source_root: Path, folder_code: str, source_id: str) -> Path:
    archive_dir = source_root / folder_code / "_archive" / source_id
    archive_dir.mkdir(parents=True, exist_ok=True)
    target = archive_dir / current_path.name
    if target.exists():
        if sha256_file(target) == sha256_file(current_path):
            return target
        target = archive_dir / f"{current_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{current_path.suffix}"
    temporary = target.with_name(f".{target.name}.tmp")
    shutil.copy2(current_path, temporary)
    os.replace(temporary, target)
    return target


def promote_download(download_path: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.name}.incoming")
    shutil.copy2(download_path, temporary)
    os.replace(temporary, destination)


def eligible(record: dict[str, Any], config: dict[str, Any], include_pending: bool) -> bool:
    if str(record.get("source_status") or "").strip().upper() not in {
        str(value).upper() for value in config["eligible_source_statuses"]
    }:
        return False
    # Collection is intentionally independent of selection.  Candidate files
    # are gathered first; operator selection only controls whether a validated
    # source enters the formal source list. Restricted sources are attempted
    # once and then remain blocked until a usable route or permission is added.
    if str(record.get("download_status") or "").strip().upper() == "PAYWALL_BLOCKED":
        return bool(str(record.get("retrieval_url") or "").strip()) and str(record.get("access_permission") or "").strip().upper() == "HIGH"
    return True


def audit_registry(config: dict[str, Any]) -> list[str]:
    workbook_path: Path = config["workbook"]
    wb = load_workbook(workbook_path, data_only=False)
    if config["sheet_name"] not in wb.sheetnames:
        return [f"Missing sheet: {config['sheet_name']}"]
    ws = wb[config["sheet_name"]]
    try:
        headers = workbook_headers(ws, int(config["header_row"]))
    except UpdaterError as exc:
        return [str(exc)]
    issues: list[str] = []
    seen_sources: set[str] = set()
    seen_snapshots: set[str] = set()
    for row in range(int(config["data_start_row"]), ws.max_row + 1):
        record = record_from_row(ws, row, headers)
        source_id = str(record.get("source_id") or "").strip()
        if not source_id:
            continue
        prefix = f"Row {row} {source_id}: "
        if source_id in seen_sources:
            issues.append(prefix + "duplicate source_id")
        seen_sources.add(source_id)
        snapshot_id = str(record.get("snapshot_id") or "").strip()
        if snapshot_id in seen_snapshots:
            issues.append(prefix + "duplicate snapshot_id")
        seen_snapshots.add(snapshot_id)
        try:
            validate_identity(record, config["folder_prefix_map"])
        except UpdaterError as exc:
            issues.append(prefix + str(exc))
            continue
        folder = config["source_root"] / str(record["folder_code"])
        if not folder.is_dir():
            issues.append(prefix + f"classification folder does not exist: {folder}")
        current_path = folder / str(record["stored_filename"])
        snapshot_status = str(record.get("snapshot_status") or "").strip().upper()
        if snapshot_status == "STORED" and not current_path.is_file():
            issues.append(prefix + f"snapshot_status=STORED but the file does not exist: {current_path}")
        if current_path.is_file() and snapshot_status != "STORED":
            issues.append(prefix + "file exists but snapshot_status is not STORED")
        recorded_hash = str(record.get("content_hash") or "").strip().lower()
        if current_path.is_file() and recorded_hash:
            actual_hash = sha256_file(current_path)
            if recorded_hash != actual_hash:
                issues.append(prefix + "content_hash does not match the current file")
    return issues


def update_one_source(
    config: dict[str, Any],
    wb,
    ws,
    headers: dict[str, int],
    row: int,
    workbook_path: Path,
    mtime_ns: int,
) -> tuple[int, str]:
    record = record_from_row(ws, row, headers)
    source_id = str(record.get("source_id") or "").strip()
    attempt_at = local_now(config)
    if (
        not str(record.get("retrieval_url") or "").strip()
        and str(record.get("access_permission") or "").strip().upper() != "HIGH"
    ):
        reason = "Official access is restricted by a paywall or licence; provide an authorised copy or retrieval route."
        set_fields(ws, row, headers, {
            "download_status": "PAYWALL_BLOCKED",
            "last_attempt_at": attempt_at,
            "content_change": "UNKNOWN",
            "failure_stage": "HTTP",
            "failure_reason": reason,
        })
        mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)
        append_log(config, {"source_id": source_id, "result": "PAYWALL_BLOCKED", "stage": "HTTP", "message": reason})
        return mtime_ns, "PAYWALL_BLOCKED"
    try:
        validate_identity(record, config["folder_prefix_map"])
    except UpdaterError as exc:
        set_fields(ws, row, headers, {
            "download_status": "FAIL",
            "last_attempt_at": attempt_at,
            "content_change": "UNKNOWN",
            "failure_stage": exc.stage,
            "failure_reason": str(exc)[:500],
        })
        mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)
        append_log(config, {"source_id": source_id, "result": "FAIL", "stage": exc.stage, "message": str(exc)})
        return mtime_ns, "FAIL"

    set_fields(ws, row, headers, {
        "last_attempt_at": attempt_at,
        "content_change": "UNKNOWN",
        "failure_stage": "NONE",
        "failure_reason": "",
    })
    mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)

    current_path = config["source_root"] / str(record["folder_code"]) / str(record["stored_filename"])
    try:
        with tempfile.TemporaryDirectory(prefix=f"source-{source_id}-") as temp_name:
            result = download_and_validate(record, config, Path(temp_name))
            current_hash = sha256_file(current_path) if current_path.is_file() else ""
            unchanged = bool(current_hash) and result.sha256 == current_hash
            success_at = local_now(config)
            if unchanged:
                set_fields(ws, row, headers, {
                    "snapshot_status": "STORED",
                    "download_status": "SUCCESS",
                    "last_success_at": success_at,
                    "content_change": "UNCHANGED",
                    "current_origin": "AUTO",
                    "failure_stage": "NONE",
                    "failure_reason": "",
                    "content_hash": result.sha256,
                })
                mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)
                append_log(config, {
                    "source_id": source_id,
                    "result": "SUCCESS",
                    "change": "UNCHANGED",
                    "bytes": result.byte_count,
                    "sha256": result.sha256,
                    "final_url": result.final_url,
                })
                return mtime_ns, "UNCHANGED"

            has_current = current_path.is_file()
            new_snapshot_id = (
                next_snapshot_id(source_id, str(record["snapshot_id"])) if has_current else str(record["snapshot_id"])
            )
            new_filename = filename_for_snapshot(record, new_snapshot_id)
            new_path = config["source_root"] / str(record["folder_code"]) / new_filename
            archived_path = None
            if has_current:
                archived_path = archive_current(
                    current_path, config["source_root"], str(record["folder_code"]), source_id
                )
            promote_download(result.path, new_path)
            stored_change = "CHANGED" if has_current else "INITIAL"
            set_fields(ws, row, headers, {
                "snapshot_id": new_snapshot_id,
                "stored_filename": new_filename,
                "snapshot_status": "STORED",
                "download_status": "SUCCESS",
                "last_success_at": success_at,
                "content_change": stored_change,
                "current_origin": "AUTO",
                "failure_stage": "NONE",
                "failure_reason": "",
                "content_hash": result.sha256,
            })
            mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)
            if has_current and current_path != new_path:
                current_path.unlink(missing_ok=True)
            append_log(config, {
                "source_id": source_id,
                "result": "SUCCESS",
                "change": stored_change,
                "snapshot_id": new_snapshot_id,
                "bytes": result.byte_count,
                "sha256": result.sha256,
                "archived_path": str(archived_path) if archived_path else None,
                "final_url": result.final_url,
            })
            return mtime_ns, stored_change
    except UpdaterError as exc:
        failure = exc
    except Exception as exc:  # unexpected failures must still preserve the current file
        failure = UpdaterError("UNKNOWN", f"Unexpected error: {type(exc).__name__}: {exc}")

    failure_status = "PAYWALL_BLOCKED" if (
        failure.stage == "HTTP"
        and str(record.get("access_permission") or "").strip().upper() != "HIGH"
        and any(marker in str(failure) for marker in ("HTTP 401", "HTTP 402", "HTTP 403"))
    ) else "FAIL"
    set_fields(ws, row, headers, {
        "download_status": failure_status,
        "last_attempt_at": attempt_at,
        "content_change": "UNKNOWN",
        "failure_stage": failure.stage,
        "failure_reason": str(failure)[:500],
    })
    mtime_ns = save_workbook_atomic(wb, workbook_path, mtime_ns)
    append_log(config, {"source_id": source_id, "result": failure_status, "stage": failure.stage, "message": str(failure)})
    return mtime_ns, failure_status


def run_updates(
    config: dict[str, Any],
    requested_ids: set[str],
    include_pending: bool,
    available_only: bool = False,
) -> int:
    workbook_path: Path = config["workbook"]
    if not workbook_path.is_file():
        raise UpdaterError("CONFIG", f"Excel workbook not found: {workbook_path}")
    with updater_lock(workbook_path, float(config["lock_stale_hours"])):
        wb = load_workbook(workbook_path, data_only=False)
        if config["sheet_name"] not in wb.sheetnames:
            raise UpdaterError("CONFIG", f"Missing sheet: {config['sheet_name']}")
        ws = wb[config["sheet_name"]]
        headers = workbook_headers(ws, int(config["header_row"]))
        candidates: list[int] = []
        for row in range(int(config["data_start_row"]), ws.max_row + 1):
            record = record_from_row(ws, row, headers)
            source_id = str(record.get("source_id") or "").strip()
            if not source_id:
                continue
            if requested_ids and source_id not in requested_ids:
                continue
            if eligible(record, config, include_pending):
                if available_only and not str(record.get("retrieval_url") or "").strip():
                    continue
                candidates.append(row)
        missing_ids = requested_ids - {
            str(record_from_row(ws, row, headers).get("source_id") or "").strip() for row in candidates
        }
        if missing_ids:
            raise UpdaterError("CONFIG", f"Requested sources do not exist or are not currently eligible: {', '.join(sorted(missing_ids))}")
        if not candidates:
            print("No collectable CURRENT sources were found.")
            return 0
        backup = create_backup(workbook_path, config["backup_root"], local_now(config))
        append_log(config, {"event": "RUN_START", "sources": len(candidates), "backup": str(backup)})
        mtime_ns = workbook_mtime(workbook_path)
        counts = {"INITIAL": 0, "CHANGED": 0, "UNCHANGED": 0, "FAIL": 0, "PAYWALL_BLOCKED": 0}
        for row in candidates:
            source_id = str(ws.cell(row=row, column=headers["source_id"]).value)
            print(f"Checking {source_id} ...", flush=True)
            mtime_ns, result = update_one_source(config, wb, ws, headers, row, workbook_path, mtime_ns)
            counts[result] += 1
            print(f"  {result}", flush=True)
        append_log(config, {"event": "RUN_END", "counts": counts})
        print(f"Completed: {counts}")
        return 1 if counts["FAIL"] else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Update source files and status fields in the Requirement Source Registry.")
    parser.add_argument("--config", default="config.json", help="Configuration file path; defaults to config.json in the current directory")
    parser.add_argument("--source-id", action="append", default=[], help="Process only the specified source_id; may be repeated")
    parser.add_argument("--include-pending", action="store_true", help="Compatibility flag; collection is independent of selection status")
    parser.add_argument("--available-only", action="store_true", help="Skip records without retrieval_url instead of recording a download failure")
    parser.add_argument("--audit-only", action="store_true", help="Audit workbook, identifiers, filenames, folders and hashes only")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config_path = Path(args.config).expanduser().resolve()
    try:
        config = read_config(config_path)
        if args.audit_only:
            issues = audit_registry(config)
            if issues:
                print("Consistency audit found issues:")
                for issue in issues:
                    print(f"- {issue}")
                return 1
            print("Consistency audit passed.")
            return 0
        return run_updates(config, set(args.source_id), args.include_pending, args.available_only)
    except UpdaterError as exc:
        print(f"Stopped [{exc.stage}]: {exc}", file=sys.stderr)
        return 2
    except FileNotFoundError as exc:
        print(f"Stopped [CONFIG]: file or folder not found: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
