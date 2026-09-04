#!/usr/bin/env python3
"""Deterministic source-maintenance stages used by the three bounded modules.

The stage boundaries intentionally match the future agent roles. No API key or
LLM is required. Each stage produces explicit, reviewable output instead of
silently making legal or certification judgments.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

import source_updater as updater


SCORE_FIELDS = (
    "authority_quality",
    "scope_relevance",
    "version_currency",
    "traceability",
    "access_permission",
    "automation_readiness",
)
VALID_SCORES = {"LOW", "MEDIUM", "HIGH"}


def read_records(config: dict[str, Any]) -> list[dict[str, Any]]:
    # Normal mode is intentional: repeated random cell access in openpyxl's
    # read-only mode reparses the XML and becomes extremely slow for this table.
    workbook = load_workbook(config["workbook"], data_only=False)
    sheet = workbook[config["sheet_name"]]
    headers = updater.workbook_headers(sheet, int(config["header_row"]))
    records: list[dict[str, Any]] = []
    for row in range(int(config["data_start_row"]), sheet.max_row + 1):
        record = updater.record_from_row(sheet, row, headers)
        if str(record.get("source_id") or "").strip():
            records.append(record)
    workbook.close()
    return records


def discovery_stage(records: list[dict[str, Any]]) -> dict[str, Any]:
    ids = [str(record["source_id"]).strip() for record in records]
    official_urls = [str(record.get("official_url") or "").strip() for record in records]
    duplicate_ids = sorted(source_id for source_id, count in Counter(ids).items() if count > 1)
    duplicate_urls = sorted(url for url, count in Counter(official_urls).items() if url and count > 1)
    no_official_url = sorted(ids[index] for index, url in enumerate(official_urls) if not url)
    provenance_gaps = sorted(
        str(record["source_id"]).strip()
        for record in records
        if not str(record.get("official_url") or "").strip()
        and (
            str(record.get("acquisition_channel") or "").strip().upper() in {"", "OFFICIAL_WEBSITE"}
            or str(record.get("provenance_status") or "").strip().upper() in {"", "UNVERIFIED"}
        )
    )
    return {
        "stage": "01_discovery",
        "purpose": "Coverage and candidate-inventory gate",
        "source_count": len(records),
        "duplicate_source_ids": duplicate_ids,
        "duplicate_official_urls": duplicate_urls,
        "no_official_url": no_official_url,
        "equivalent_provenance_gaps": provenance_gaps,
        "status": "PASS" if not duplicate_ids and not provenance_gaps else "REVIEW",
    }


def monitoring_stage(records: list[dict[str, Any]], config: dict[str, Any], include_pending: bool) -> dict[str, Any]:
    eligible_ids: list[str] = []
    downloadable_ids: list[str] = []
    missing_retrieval_ids: list[str] = []
    update_models: Counter[str] = Counter()
    acquisition_channels: Counter[str] = Counter()
    for record in records:
        update_models[str(record.get("update_model") or "").strip().upper()] += 1
        acquisition_channels[str(record.get("acquisition_channel") or "").strip().upper()] += 1
        if not updater.eligible(record, config, include_pending):
            continue
        source_id = str(record["source_id"]).strip()
        eligible_ids.append(source_id)
        if str(record.get("retrieval_url") or "").strip():
            downloadable_ids.append(source_id)
        else:
            missing_retrieval_ids.append(source_id)
    return {
        "stage": "02_update_monitor",
        "purpose": "Determine which sources can be checked in this run",
        "eligible_count": len(eligible_ids),
        "downloadable_count": len(downloadable_ids),
        "missing_retrieval_url": missing_retrieval_ids,
        "update_model_counts": dict(update_models),
        "acquisition_channel_counts": dict(acquisition_channels),
        "status": "PASS",
    }


def assessment_stage(records: list[dict[str, Any]]) -> dict[str, Any]:
    issues: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    for record in records:
        source_id = str(record["source_id"]).strip()
        if not str(record.get("inclusion_rationale") or "").strip():
            issues.append({"source_id": source_id, "field": "inclusion_rationale", "issue": "missing"})
        for field in SCORE_FIELDS:
            score = str(record.get(field) or "").strip().upper()
            if score not in VALID_SCORES:
                issues.append({"source_id": source_id, "field": field, "issue": "missing_or_invalid"})
        document_type = str(record.get("document_type") or "").strip()
        requirement_role = str(record.get("requirement_role") or "").strip()
        applicability = str(record.get("applicability_reference") or "").strip()
        if document_type == "Equipment manual / OEM instruction" or requirement_role == "Asset-specific normative":
            if str(record.get("source_family") or "").strip() != "Manufacturer / supplier":
                issues.append({"source_id": source_id, "field": "source_family", "issue": "equipment_source_not_manufacturer"})
            if not applicability:
                issues.append({"source_id": source_id, "field": "applicability_reference", "issue": "equipment_applicability_missing"})
        status_counts[updater.selection_from_scores(record)] += 1
    return {
        "stage": "03_selection_review",
        "purpose": "Verify that the initial source-quality assessment is complete",
        "selection_status_counts": dict(status_counts),
        "issues": issues,
        "status": "PASS" if not issues else "REVIEW",
    }


def collection_stage(
    config: dict[str, Any],
    include_pending: bool,
    available_only: bool,
    execute: bool,
) -> dict[str, Any]:
    if not execute:
        return {
            "stage": "04_collection",
            "purpose": "Download, validate, name, archive and store original-format snapshots",
            "executed": False,
            "status": "DRY_RUN",
        }
    result = updater.run_updates(config, set(), include_pending, available_only)
    return {
        "stage": "04_collection",
        "purpose": "Download, validate, name, archive and store original-format snapshots",
        "executed": True,
        "updater_exit_code": result,
        "status": "PASS" if result == 0 else "REVIEW",
    }


def metadata_stage(config: dict[str, Any]) -> dict[str, Any]:
    issues = updater.audit_registry(config)
    records = read_records(config)
    download_counts = Counter(str(record.get("download_status") or "").strip() for record in records)
    stored_count = sum(str(record.get("snapshot_status") or "").strip() == "STORED" for record in records)
    return {
        "stage": "05_metadata_completion",
        "purpose": "Verify workbook-to-file consistency and summarize final state",
        "stored_count": stored_count,
        "download_status_counts": dict(download_counts),
        "audit_issues": issues,
        "status": "PASS" if not issues else "REVIEW",
    }


def write_report(config: dict[str, Any], report: dict[str, Any]) -> Path:
    output_dir: Path = config["log_root"] / "pipeline_reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"pipeline_{datetime.now():%Y%m%d_%H%M%S}.json"
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def run_pipeline(config_path: Path, include_pending: bool, available_only: bool, execute: bool) -> int:
    config = updater.read_config(config_path)
    initial_records = read_records(config)
    stages = [
        discovery_stage(initial_records),
        monitoring_stage(initial_records, config, include_pending),
        assessment_stage(initial_records),
    ]
    try:
        stages.append(collection_stage(config, include_pending, available_only, execute))
    except updater.UpdaterError as exc:
        stages.append({
            "stage": "04_collection",
            "purpose": "Download, validate, name, archive and store original-format snapshots",
            "executed": execute,
            "status": "STOP",
            "failure_stage": exc.stage,
            "failure_reason": str(exc),
        })
    stages.append(metadata_stage(config))
    report = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "config": str(config_path),
        "stages": stages,
        "overall_status": "PASS" if all(stage["status"] in {"PASS", "DRY_RUN"} for stage in stages) else "REVIEW",
    }
    output = write_report(config, report)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"Report: {output}")
    return 0 if report["overall_status"] == "PASS" else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the deterministic Requirement source pipeline stages.")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--include-pending", action="store_true")
    parser.add_argument("--available-only", action="store_true")
    parser.add_argument("--execute", action="store_true", help="Run downloads; without this flag the collection stage is a dry run")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    return run_pipeline(Path(args.config).expanduser().resolve(), args.include_pending, args.available_only, args.execute)


if __name__ == "__main__":
    raise SystemExit(main())
