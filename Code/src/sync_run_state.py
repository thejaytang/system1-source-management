#!/usr/bin/env python3
"""Safely merge program-owned fields from a staging run into the live workbook."""

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

import source_updater as updater


IDENTITY_GUARDS = ("source_title", "retrieval_url", "folder_code", "file_format")


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def rows_by_source(ws, headers: dict[str, int], data_start_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(data_start_row, ws.max_row + 1):
        source_id = str(ws.cell(row=row, column=headers["source_id"]).value or "").strip()
        if not source_id:
            continue
        if source_id in rows:
            raise updater.UpdaterError("CONFIG", f"Duplicate source_id: {source_id}")
        rows[source_id] = row
    return rows


def manual_is_newer(live: dict[str, Any], staged: dict[str, Any]) -> bool:
    if str(live.get("current_origin") or "").strip().upper() != "MANUAL":
        return False
    manual_date = as_date(live.get("manual_update_date"))
    attempt_date = as_date(staged.get("last_attempt_at"))
    if manual_date is None:
        return True
    return attempt_date is None or manual_date >= attempt_date


def sync_state(config: dict[str, Any], staging_workbook: Path) -> dict[str, Any]:
    live_path: Path = config["workbook"]
    if not live_path.is_file():
        raise updater.UpdaterError("CONFIG", f"Live Excel workbook not found: {live_path}")
    if not staging_workbook.is_file():
        raise updater.UpdaterError("CONFIG", f"Staging Excel workbook not found: {staging_workbook}")

    with updater.updater_lock(live_path, float(config["lock_stale_hours"])):
        live_wb = load_workbook(live_path, data_only=False)
        staged_wb = load_workbook(staging_workbook, data_only=False)
        live_ws = live_wb[config["sheet_name"]]
        staged_ws = staged_wb[config["sheet_name"]]
        live_headers = updater.workbook_headers(live_ws, int(config["header_row"]))
        staged_headers = updater.workbook_headers(staged_ws, int(config["header_row"]))
        live_rows = rows_by_source(live_ws, live_headers, int(config["data_start_row"]))
        staged_rows = rows_by_source(staged_ws, staged_headers, int(config["data_start_row"]))

        synced: list[str] = []
        skipped_not_run: list[str] = []
        conflicts: list[dict[str, str]] = []
        for source_id, staged_row in staged_rows.items():
            if source_id not in live_rows:
                conflicts.append({"source_id": source_id, "reason": "missing_from_live_workbook"})
                continue
            live_record = updater.record_from_row(live_ws, live_rows[source_id], live_headers)
            staged_record = updater.record_from_row(staged_ws, staged_row, staged_headers)
            if str(staged_record.get("download_status") or "").strip().upper() == "NOT_RUN":
                skipped_not_run.append(source_id)
                continue
            changed_identity = [
                field for field in IDENTITY_GUARDS
                if str(live_record.get(field) or "").strip() != str(staged_record.get(field) or "").strip()
            ]
            if changed_identity:
                conflicts.append({"source_id": source_id, "reason": "identity_changed:" + ",".join(changed_identity)})
                continue
            if manual_is_newer(live_record, staged_record):
                conflicts.append({"source_id": source_id, "reason": "newer_manual_snapshot_preserved"})
                continue
            updates = {field: staged_record.get(field) for field in updater.PROGRAM_FIELDS}
            updater.set_fields(live_ws, live_rows[source_id], live_headers, updates)
            synced.append(source_id)

        staged_wb.close()
        backup = updater.create_backup(live_path, config["backup_root"], updater.local_now(config))
        mtime_ns = updater.workbook_mtime(live_path)
        updater.save_workbook_atomic(live_wb, live_path, mtime_ns)
        result = {
            "synced": synced,
            "skipped_not_run": skipped_not_run,
            "conflicts": conflicts,
            "backup": str(backup),
        }
        updater.append_log(config, {
            "event": "STAGING_STATE_SYNC",
            "synced_count": len(synced),
            "skipped_not_run_count": len(skipped_not_run),
            "conflict_count": len(conflicts),
            "backup": str(backup),
        })
        return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Merge staging download state into the live source registry.")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--staging-workbook", required=True)
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config = updater.read_config(Path(args.config).expanduser().resolve())
    try:
        result = sync_state(config, Path(args.staging_workbook).expanduser().resolve())
    except updater.UpdaterError as exc:
        print(f"Stopped [{exc.stage}]: {exc}")
        return 2
    print(f"Sync completed: {len(result['synced'])}; skipped NOT_RUN: {len(result['skipped_not_run'])}; conflicts: {len(result['conflicts'])}")
    for conflict in result["conflicts"]:
        print(f"- {conflict['source_id']}: {conflict['reason']}")
    return 1 if result["conflicts"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
