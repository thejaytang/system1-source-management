from __future__ import annotations

import json
import random
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import human_operations as operations


SOURCE_HEADERS = [
    "source_id", "snapshot_id", "stored_filename", "snapshot_status", "selection_status",
    "download_status", "last_attempt_at", "last_success_at", "content_change", "current_origin",
    "current_snapshot_date", "failure_stage", "failure_reason", "needs_human_action", "content_hash",
    "source_title", "official_url", "retrieval_url", "folder_code", "file_format", "issuer",
    "jurisdiction", "authoritative_language", "version", "effective_date", "source_family",
    "document_type", "requirement_role", "inclusion_rationale", "authority_quality", "scope_relevance",
    "version_currency", "traceability", "access_permission", "automation_readiness",
    "operator_selection_decision", "manual_update_date", "manual_updated_by", "source_status",
    "primary_source_id", "notes", "acquisition_channel", "provenance_status", "update_model",
    "applicability_reference",
]


class HumanOperationsTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name) / "Requirement Source Management"
        self.program = self.root / "Code"
        self.config_dir = self.program / "config"
        self.config_dir.mkdir(parents=True)
        (self.program / "runtime" / "inbox").mkdir(parents=True)
        self.workbook_path = self.root / "Requirement_Source_Registry.xlsx"
        self.config_path = self.config_dir / "config.json"
        self._build_workbook()
        self.config_path.write_text(json.dumps({
            "workbook": "../../Requirement_Source_Registry.xlsx",
            "source_root": "../../Data",
            "backup_root": "../runtime/backups",
            "log_root": "../runtime/logs",
            "human_operation_sheet": "Human Operation Desktop",
            "discovery_candidate_inbox": "../runtime/inbox/discovery_candidates.json",
            "folder_prefix_map": {"A_Public_Authority": "PA"},
            "random_qa": {"enabled": True, "schedule": "LAST_DAY_OF_MONTH", "monthly_sample_size": 1, "max_sample_size": 20},
        }), encoding="utf-8")
        self.inbox = self.program / "runtime" / "inbox" / "discovery_candidates.json"
        self.inbox.write_text("[]", encoding="utf-8")

    def tearDown(self):
        self.temp.cleanup()

    def _build_workbook(self):
        wb = Workbook()
        source = wb.active
        source.title = "Source Register"
        source.append(["GROUP"] * len(SOURCE_HEADERS))
        source.append(SOURCE_HEADERS)
        values = {name: None for name in SOURCE_HEADERS}
        values.update({
            "source_id": "PA001", "snapshot_id": "PA001-001", "source_title": "Existing law",
            "official_url": "https://example.test/law", "retrieval_url": "https://example.test/law",
            "folder_code": "A_Public_Authority", "file_format": "html",
            "stored_filename": "PA001-001_Existing_law.html", "snapshot_status": "STORED",
            "issuer": "Test authority", "jurisdiction": "Norway", "authoritative_language": "English",
            "source_family": "Public authority", "document_type": "Law", "requirement_role": "Primary normative",
            "inclusion_rationale": "Test source", "authority_quality": "HIGH", "scope_relevance": "HIGH",
            "version_currency": "HIGH", "traceability": "HIGH", "access_permission": "HIGH",
            "automation_readiness": "HIGH", "operator_selection_decision": "INCLUDE",
            "download_status": "SUCCESS", "content_change": "UNCHANGED",
            "current_origin": "AUTO", "failure_stage": "NONE", "source_status": "CURRENT",
            "acquisition_channel": "OFFICIAL_WEBSITE", "provenance_status": "OFFICIAL_ORIGINAL", "update_model": "ROLLING",
        })
        source.append([values[name] for name in SOURCE_HEADERS])
        source["E3"] = '=IF(AJ3="EXCLUDE","EXCLUDE",IF(AND(AJ3="INCLUDE",F3="SUCCESS",D3="STORED"),"INCLUDE","PENDING"))'
        source["K3"] = '=IF(J3="AUTO",H3,IF(J3="MANUAL",AK3,""))'
        source["N3"] = '=IF(OR(F3="FAIL",F3="PAYWALL_BLOCKED"),"YES","NO")'
        ops = wb.create_sheet("Human Operation Desktop")
        ops.append(["OPERATION"] * len(operations.HUMAN_HEADERS))
        ops.append(list(operations.HUMAN_HEADERS))
        dashboard = wb.create_sheet("Dashboard")
        dashboard["A46"] = "NO"
        dashboard["E46"] = 1
        wb.save(self.workbook_path)

    def _candidate(self):
        return {
            "source_title": "New official regulation",
            "official_url": "https://example.test/new-regulation",
            "retrieval_url": "https://example.test/new-regulation",
            "folder_code": "A_Public_Authority",
            "file_format": "html",
            "issuer": "Test authority",
            "jurisdiction": "Norway",
            "authoritative_language": "English",
            "source_family": "Public authority",
            "document_type": "Regulation",
            "requirement_role": "Primary normative",
            "inclusion_rationale": "Potentially relevant new regulation requiring human gate approval.",
        }

    def test_discovery_candidate_requires_accept_before_source_register(self):
        self.inbox.write_text(json.dumps([self._candidate()]), encoding="utf-8")
        result = operations.sync_discovery_candidates(self.config_path)
        self.assertEqual(result["added_count"], 1)
        before = load_workbook(self.workbook_path, data_only=False)
        self.assertEqual(before["Source Register"].max_row, 3)
        operation = before["Human Operation Desktop"]
        self.assertEqual(operation["B3"].value, "NEW_SOURCE_CANDIDATE")
        before.close()

        wb = load_workbook(self.workbook_path)
        operation = wb["Human Operation Desktop"]
        operation["G3"] = "ACCEPT"
        operation["H3"] = "Reviewer"
        operation["J3"] = "Confirmed official candidate."
        wb.save(self.workbook_path)
        operations.process_decisions(self.config_path)

        after = load_workbook(self.workbook_path, data_only=False)
        self.assertEqual(after["Source Register"]["A4"].value, "PA002")
        self.assertEqual(after["Human Operation Desktop"]["K3"].value, "APPLIED")
        self.assertTrue(after["Human Operation Desktop"].row_dimensions[3].hidden)
        after.close()

    def test_rejected_candidate_is_archived_without_register_row(self):
        self.inbox.write_text(json.dumps([self._candidate()]), encoding="utf-8")
        operations.sync_discovery_candidates(self.config_path)
        wb = load_workbook(self.workbook_path)
        ops = wb["Human Operation Desktop"]
        ops["G3"] = "REJECT"
        ops["H3"] = "Reviewer"
        ops["J3"] = "Outside approved scope."
        wb.save(self.workbook_path)
        operations.process_decisions(self.config_path)
        after = load_workbook(self.workbook_path)
        self.assertEqual(after["Source Register"].max_row, 3)
        self.assertEqual(after["Human Operation Desktop"]["K3"].value, "REJECTED")
        self.assertTrue(after["Human Operation Desktop"].row_dimensions[3].hidden)
        after.close()

    def test_random_qa_request_creates_and_archives_check(self):
        result = operations.generate_random_qa(self.config_path, random.Random(1), today=date(2026, 8, 31))
        self.assertEqual(result["generated_count"], 1)
        wb = load_workbook(self.workbook_path)
        ops = wb["Human Operation Desktop"]
        self.assertEqual(ops["B3"].value, "RANDOM_QA_CHECK")
        ops["G3"] = "CORRECT"
        ops["H3"] = "Reviewer"
        wb.save(self.workbook_path)
        operations.process_decisions(self.config_path)
        after = load_workbook(self.workbook_path)
        self.assertEqual(after["Human Operation Desktop"]["K3"].value, "APPLIED")
        self.assertIsNotNone(after["Human Operation Desktop"]["L3"].value)
        self.assertTrue(after["Human Operation Desktop"].row_dimensions[3].hidden)
        self.assertNotIn("QA Inspection Log", after.sheetnames)
        first_operated_at = after["Human Operation Desktop"]["L3"].value
        after.close()

        operations.process_decisions(self.config_path)
        repeated = load_workbook(self.workbook_path)
        self.assertEqual(repeated["Human Operation Desktop"].max_row, 3)
        self.assertEqual(repeated["Human Operation Desktop"]["L3"].value, first_operated_at)
        repeated.close()

        duplicate = operations.generate_random_qa(self.config_path, random.Random(1), today=date(2026, 8, 31))
        self.assertEqual(duplicate["status"], "ALREADY_CREATED")
        self.assertEqual(duplicate["generated_count"], 0)

    def test_random_qa_is_not_created_before_month_end(self):
        result = operations.generate_random_qa(self.config_path, random.Random(1), today=date(2026, 8, 30))
        self.assertEqual(result["status"], "NOT_DUE")
        wb = load_workbook(self.workbook_path)
        self.assertEqual(wb["Human Operation Desktop"].max_row, 2)
        wb.close()

    def test_review_queue_action_is_applied_without_manual_source_row_edit(self):
        item = {
            "review_id": "abc123",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "ROLLING_CONTENT_CHANGED",
            "reason": "Content changed.",
            "recommended_action": "Review impact.",
        }
        operations.sync_review_queue(self.config_path, [item])
        wb = load_workbook(self.workbook_path)
        ops = wb["Human Operation Desktop"]
        ops["G3"] = "ACCEPT"
        ops["H3"] = "Reviewer"
        ops["J3"] = "No downstream requirement change."
        wb.save(self.workbook_path)
        operations.process_decisions(self.config_path)
        after = load_workbook(self.workbook_path)
        source = after["Source Register"]
        headers = {cell.value: cell.column for cell in source[2] if cell.value}
        self.assertEqual(source.cell(3, headers["manual_updated_by"]).value, "Reviewer")
        self.assertIn("No downstream requirement change", source.cell(3, headers["notes"]).value)
        self.assertTrue(after["Human Operation Desktop"].row_dimensions[3].hidden)
        after.close()

    def test_selection_review_applies_human_fields_from_desktop(self):
        item = {
            "review_id": "selection123",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "SELECTION_PENDING",
            "reason": "Scope requires confirmation.",
            "recommended_action": "Review selection.",
            "review_fingerprint": "abc123",
        }
        operations.sync_review_queue(self.config_path, [item])
        wb = load_workbook(self.workbook_path)
        ops = wb["Human Operation Desktop"]
        headers = {cell.value: cell.column for cell in ops[2] if cell.value}
        ops.cell(3, headers["scope_relevance"], "MEDIUM")
        ops.cell(3, headers["operator_selection_decision"], "PENDING")
        ops.cell(3, headers["decision"], "ACCEPT")
        ops.cell(3, headers["operator"], "Reviewer")
        ops.cell(3, headers["operator_note"], "Keep pending until site scope is confirmed.")
        wb.save(self.workbook_path)
        operations.process_decisions(self.config_path)
        after = load_workbook(self.workbook_path)
        source = after["Source Register"]
        source_headers = {cell.value: cell.column for cell in source[2] if cell.value}
        self.assertEqual(source.cell(3, source_headers["scope_relevance"]).value, "MEDIUM")
        self.assertEqual(source.cell(3, source_headers["operator_selection_decision"]).value, "PENDING")
        self.assertEqual(after["Human Operation Desktop"]["K3"].value, "APPLIED")
        self.assertTrue(after["Human Operation Desktop"].row_dimensions[3].hidden)
        after.close()

    def test_resolved_download_review_is_hidden_without_human_decision(self):
        item = {
            "review_id": "download123",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "DOWNLOAD_FAILURE",
            "reason": "Earlier retrieval failed.",
            "recommended_action": "Retry the source.",
        }
        operations.sync_review_queue(self.config_path, [item])
        result = operations.reconcile_resolved_reviews(self.config_path)
        self.assertEqual(result["resolved_count"], 1)
        after = load_workbook(self.workbook_path)
        ops = after["Human Operation Desktop"]
        self.assertEqual(ops["G3"].value, "PENDING")
        self.assertEqual(ops["K3"].value, "APPLIED")
        self.assertIsNotNone(ops["L3"].value)
        self.assertTrue(ops.row_dimensions[3].hidden)
        after.close()

    def test_stale_review_does_not_overwrite_newer_source_state(self):
        item = {
            "review_id": "stale123",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "SELECTION_PENDING",
            "reason": "Scope requires confirmation.",
            "recommended_action": "Review selection.",
        }
        operations.sync_review_queue(self.config_path, [item])
        workbook = load_workbook(self.workbook_path)
        source = workbook["Source Register"]
        source_headers = {cell.value: cell.column for cell in source[2] if cell.value}
        source.cell(3, source_headers["notes"], "A newer source edit")
        ops = workbook["Human Operation Desktop"]
        headers = {cell.value: cell.column for cell in ops[2] if cell.value}
        ops.cell(3, headers["decision"], "APPLY")
        ops.cell(3, headers["operator"], "Reviewer")
        workbook.save(self.workbook_path)
        operations.process_decisions(self.config_path)
        after = load_workbook(self.workbook_path)
        self.assertEqual(after["Human Operation Desktop"]["K3"].value, "WAITING_FOR_HUMAN")
        self.assertIn("changed after this task", after["Human Operation Desktop"]["M3"].value)
        after.close()

    def test_dashboard_priority_items_use_new_range_and_priority_order(self):
        workbook = load_workbook(self.workbook_path)
        desktop = workbook["Human Operation Desktop"]
        headers = operations.operation_headers(desktop)
        today = date.today()
        tasks = [
            ("OP-1", "PA001", "Pending old", "PENDING", today.replace(year=today.year - 1)),
            ("OP-2", "PA002", "Waiting recent", "WAITING_FOR_HUMAN", today),
            ("OP-3", "PA003", "Replan recent", "NEEDS_REPLAN", today),
            ("OP-4", "PA004", "Replan old", "NEEDS_REPLAN", today.replace(year=today.year - 1)),
            ("OP-5", "PA005", "Pending recent", "PENDING", today),
            ("OP-6", "PA006", "Waiting old", "WAITING_FOR_HUMAN", today.replace(year=today.year - 1)),
        ]
        for operation_id, source_id, title, status, created_at in tasks:
            operations.append_operation(desktop, headers, {
                "operation_id": operation_id,
                "operation_type": "SOURCE_REVIEW",
                "source_id": source_id,
                "source_title": title,
                "trigger": "TEST_ISSUE",
                "proposed_action": "Review the source.",
                "decision": "PENDING",
                "program_status": status,
                "created_at": created_at,
            })
        operations.prepare_human_workbook(workbook, desktop, headers)

        dashboard = workbook["Dashboard"]
        self.assertEqual(
            [dashboard.cell(row, 1).value for row in range(36, 41)],
            ["PA004", "PA003", "PA006", "PA002", "PA001"],
        )
        self.assertGreater(dashboard["E36"].value, dashboard["E37"].value)
        self.assertEqual(dashboard["F36"].value, "NEEDS_REPLAN")
        self.assertEqual(dashboard["F38"].value, "WAITING_FOR_HUMAN")
        self.assertIsNone(dashboard["A17"].value)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
