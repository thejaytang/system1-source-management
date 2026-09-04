from __future__ import annotations

import hashlib
import io
import json
import random
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import human_operations
import leader_orchestrator
import manual_intake
import source_updater
from system1.cli import run_cycle


SOURCE_HEADERS = [
    "source_id", "snapshot_id", "stored_filename", "snapshot_status", "selection_status",
    "download_status", "last_attempt_at", "last_success_at", "content_change", "current_origin",
    "current_snapshot_date", "failure_stage", "failure_reason", "needs_human_action", "content_hash",
    "source_title", "official_url", "retrieval_url", "folder_code", "file_format", "issuer",
    "jurisdiction", "authoritative_language", "version", "effective_date", "source_family",
    "document_type", "requirement_role", "inclusion_rationale", "authority_quality", "scope_relevance",
    "version_currency", "traceability", "access_permission", "automation_readiness",
    "operator_selection_decision", "manual_update_date", "manual_updated_by", "source_status",
    "primary_source_id", "notes", "acquisition_channel", "provenance_status", "update_model",
    "applicability_reference",
]


def html_document(title: str, official_url: str, marker: str = "current") -> bytes:
    body = (
        f"<!doctype html><html><head><title>{title}</title></head><body>"
        f"<h1>{title}</h1><p>Official source: <a href=\"{official_url}\">{official_url}</a></p>"
        f"<p>{marker} aquaculture compliance requirements and official guidance. "
        "This paragraph makes the test document large enough for content validation.</p>"
        "</body></html>"
    )
    return body.encode("utf-8")


class OperatorJourneyTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name) / "System1 Test Workspace With Spaces"
        self.code = self.root / "Code"
        self.config_dir = self.code / "config"
        self.runtime = self.code / "runtime"
        self.data = self.root / "Data"
        self.intake = self.data / "00_Human_Intake"
        self.authority = self.data / "A_Public_Authority"
        self.config_dir.mkdir(parents=True)
        (self.runtime / "inbox").mkdir(parents=True)
        self.intake.mkdir(parents=True)
        self.authority.mkdir(parents=True)
        self.workbook = self.root / "Requirement_Source_Registry.xlsx"
        self.config_path = self.config_dir / "config.json"
        self.inbox = self.runtime / "inbox" / "discovery_candidates.json"
        self.inbox.write_text("[]", encoding="utf-8")
        self.current_url = "https://www.regjeringen.no/en/existing-law/"
        self.current_name = "PA001-001_Existing_law.html"
        self.current_bytes = html_document("Existing law", self.current_url)
        (self.authority / self.current_name).write_bytes(self.current_bytes)
        self._build_workbook()
        self.config_path.write_text(json.dumps({
            "workbook": "../../Requirement_Source_Registry.xlsx",
            "source_root": "../../Data",
            "backup_root": "../runtime/backups",
            "log_root": "../runtime/logs",
            "manual_intake_root": "../../Data/00_Human_Intake",
            "manual_intake_archive_root": "../runtime/manual_intake_archive",
            "sheet_name": "Source Register",
            "human_operation_sheet": "Human Operation Desktop",
            "discovery_candidate_inbox": "../runtime/inbox/discovery_candidates.json",
            "header_row": 2,
            "data_start_row": 3,
            "timezone": "Europe/Oslo",
            "minimum_pdf_bytes": 100,
            "minimum_html_bytes": 100,
            "minimum_binary_bytes": 100,
            "format_preference": ["html", "pdf", "xlsx", "zip"],
            "eligible_source_statuses": ["CURRENT"],
            "folder_prefix_map": {"A_Public_Authority": "PA", "Z_Pending_Classification": "PC"},
            "lock_stale_hours": 12,
            "random_qa": {
                "enabled": True,
                "schedule": "LAST_DAY_OF_MONTH",
                "monthly_sample_size": 1,
                "max_sample_size": 20,
            },
            "leader": {
                "review_threshold_days": 60,
                "module_interval_days": {
                    "discovery_and_intake": 0,
                    "retrieval_and_monitoring": 0,
                    "governance_and_qa": 0,
                },
            },
        }), encoding="utf-8")

    def tearDown(self):
        self.temporary.cleanup()

    def _build_workbook(self):
        workbook = Workbook()
        source = workbook.active
        source.title = "Source Register"
        source.append(["GROUP"] * len(SOURCE_HEADERS))
        source.append(SOURCE_HEADERS)
        values = {name: None for name in SOURCE_HEADERS}
        values.update({
            "source_id": "PA001",
            "snapshot_id": "PA001-001",
            "stored_filename": self.current_name,
            "snapshot_status": "STORED",
            "download_status": "SUCCESS",
            "content_change": "UNCHANGED",
            "current_origin": "AUTO",
            "failure_stage": "NONE",
            "content_hash": hashlib.sha256(self.current_bytes).hexdigest(),
            "source_title": "Existing law",
            "official_url": self.current_url,
            "retrieval_url": self.current_url,
            "folder_code": "A_Public_Authority",
            "file_format": "html",
            "issuer": "Norwegian Government",
            "jurisdiction": "Norway",
            "authoritative_language": "English",
            "source_family": "Public authority",
            "document_type": "Law",
            "requirement_role": "Primary normative",
            "inclusion_rationale": "Official law relevant to the controlled source scope.",
            "authority_quality": "HIGH",
            "scope_relevance": "HIGH",
            "version_currency": "HIGH",
            "traceability": "HIGH",
            "access_permission": "HIGH",
            "automation_readiness": "HIGH",
            "operator_selection_decision": "INCLUDE",
            "source_status": "CURRENT",
            "notes": "Initial record.",
            "acquisition_channel": "OFFICIAL_WEBSITE",
            "provenance_status": "OFFICIAL_ORIGINAL",
            "update_model": "ROLLING",
        })
        source.append([values[name] for name in SOURCE_HEADERS])
        source["E3"] = '=IF(AJ3="EXCLUDE","EXCLUDE",IF(AND(AJ3="INCLUDE",F3="SUCCESS",D3="STORED"),"INCLUDE","PENDING"))'
        source["K3"] = '=IF(J3="AUTO",H3,IF(J3="MANUAL",AK3,""))'
        source["N3"] = '=IF(OR(F3="FAIL",F3="PAYWALL_BLOCKED"),"YES","NO")'
        operations = workbook.create_sheet("Human Operation Desktop")
        operations.append(["OPERATION"] * len(human_operations.HUMAN_HEADERS))
        operations.append(list(human_operations.HUMAN_HEADERS))
        dashboard = workbook.create_sheet("Dashboard")
        dashboard["A46"] = "NO"
        dashboard["E46"] = 1
        workbook.save(self.workbook)

    def _operation_headers(self, sheet):
        return human_operations.operation_headers(sheet)

    def _operation_rows(self, *, operation_type: str | None = None, trigger: str | None = None) -> list[int]:
        workbook = load_workbook(self.workbook)
        sheet = workbook["Human Operation Desktop"]
        headers = self._operation_headers(sheet)
        rows = []
        for row in range(3, sheet.max_row + 1):
            if operation_type and str(sheet.cell(row, headers["operation_type"]).value or "") != operation_type:
                continue
            if trigger and trigger not in str(sheet.cell(row, headers["trigger"]).value or ""):
                continue
            if sheet.cell(row, headers["operation_id"]).value:
                rows.append(row)
        workbook.close()
        return rows

    def _set_operation(
        self,
        row: int,
        action: str,
        operator: str = "Ana",
        note: str = "Reviewed against the official source.",
        **fields,
    ) -> None:
        workbook = load_workbook(self.workbook)
        sheet = workbook["Human Operation Desktop"]
        headers = self._operation_headers(sheet)
        sheet.cell(row, headers["decision"], action)
        sheet.cell(row, headers["operator"], operator)
        sheet.cell(row, headers["operator_note"], note)
        for name, value in fields.items():
            sheet.cell(row, headers[name], value)
        workbook.save(self.workbook)
        workbook.close()

    def _source_record(self, source_id: str = "PA001") -> dict:
        workbook = load_workbook(self.workbook, data_only=False)
        sheet = workbook["Source Register"]
        headers = source_updater.workbook_headers(sheet, 2)
        row = human_operations.find_source_row(sheet, headers, source_id)
        record = source_updater.record_from_row(sheet, row, headers)
        workbook.close()
        return record

    def test_existing_source_review_applies_note_and_return_requires_replan(self):
        review = {
            "review_id": "impact-review",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "ROLLING_CONTENT_CHANGED",
            "reason": "The official page changed.",
            "recommended_action": "Review the downstream impact.",
        }
        human_operations.sync_review_queue(self.config_path, [review])
        row = self._operation_rows(trigger="ROLLING_CONTENT_CHANGED")[0]
        self._set_operation(row, "APPLY", note="No controlled Requirement changed.")
        result = human_operations.process_decisions(self.config_path)
        self.assertEqual(result["processed_count"], 1)
        record = self._source_record()
        self.assertEqual(record["manual_updated_by"], "Ana")
        self.assertIn("No controlled Requirement changed", record["notes"])
        workbook = load_workbook(self.workbook)
        operations = workbook["Human Operation Desktop"]
        headers = self._operation_headers(operations)
        self.assertEqual(operations.cell(row, headers["program_status"]).value, "APPLIED")
        self.assertIsNotNone(operations.cell(row, headers["checked_at"]).value)
        self.assertIsNotNone(operations.cell(row, headers["program_operated_at"]).value)
        self.assertTrue(bool(operations.row_dimensions[row].hidden))
        workbook.close()

        returned = dict(review, review_id="provenance-review", trigger="PROVENANCE_UNVERIFIED")
        human_operations.sync_review_queue(self.config_path, [returned])
        row = self._operation_rows(trigger="PROVENANCE_UNVERIFIED")[0]
        self._set_operation(row, "RETURN", note="The proposed provenance is not supported.")
        human_operations.process_decisions(self.config_path)
        workbook = load_workbook(self.workbook)
        sheet = workbook["Human Operation Desktop"]
        headers = self._operation_headers(sheet)
        self.assertEqual(sheet.cell(row, headers["program_status"]).value, "NEEDS_REPLAN")
        self.assertFalse(bool(sheet.row_dimensions[row].hidden))
        workbook.close()

    def test_selection_review_handles_pending_note_and_effective_include_gate(self):
        review = {
            "review_id": "selection-review",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "SELECTION_PENDING",
            "reason": "Scope needs confirmation.",
            "recommended_action": "Choose INCLUDE, PENDING or EXCLUDE.",
        }
        human_operations.sync_review_queue(self.config_path, [review])
        row = self._operation_rows(operation_type="SELECTION_REVIEW")[0]
        self._set_operation(row, "APPLY", note="", operator_selection_decision="PENDING")
        waiting = human_operations.process_decisions(self.config_path)
        self.assertEqual(waiting["waiting_count"], 1)

        self._set_operation(
            row, "APPLY", note="Site applicability evidence is still missing.",
            operator_selection_decision="PENDING", scope_relevance="MEDIUM",
        )
        processed = human_operations.process_decisions(self.config_path)
        self.assertEqual(processed["processed_count"], 1)
        self.assertEqual(self._source_record()["operator_selection_decision"], "PENDING")

        record = self._source_record()
        record["operator_selection_decision"] = "INCLUDE"
        record["download_status"] = "FAIL"
        self.assertEqual(source_updater.selection_from_scores(record), "PENDING")

    def test_selection_review_can_record_exclude_then_restore_include_intent(self):
        review = {
            "review_id": "selection-exclude",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "SELECTION_PENDING",
            "reason": "Selection must be confirmed.",
            "recommended_action": "Choose INCLUDE, PENDING or EXCLUDE.",
        }
        human_operations.sync_review_queue(self.config_path, [review])
        row = self._operation_rows(operation_type="SELECTION_REVIEW")[0]
        self._set_operation(
            row, "APPLY", note="Not applicable to the current system boundary.",
            operator_selection_decision="EXCLUDE",
        )
        human_operations.process_decisions(self.config_path)
        excluded = self._source_record()
        self.assertEqual(excluded["operator_selection_decision"], "EXCLUDE")
        self.assertEqual(source_updater.selection_from_scores(excluded), "EXCLUDE")

        review["review_id"] = "selection-include"
        human_operations.sync_review_queue(self.config_path, [review])
        row = self._operation_rows(operation_type="SELECTION_REVIEW")[-1]
        self._set_operation(
            row, "APPLY", note="Applicability and all governance gates are now confirmed.",
            operator_selection_decision="INCLUDE",
        )
        human_operations.process_decisions(self.config_path)
        included = self._source_record()
        self.assertEqual(included["operator_selection_decision"], "INCLUDE")
        self.assertEqual(source_updater.selection_from_scores(included), "INCLUDE")

    def test_new_source_candidate_accept_and_reject_are_human_gated(self):
        candidate = {
            "source_title": "New official regulation",
            "official_url": "https://www.regjeringen.no/en/new-regulation/",
            "retrieval_url": "https://www.regjeringen.no/en/new-regulation/",
            "folder_code": "A_Public_Authority",
            "file_format": "html",
            "issuer": "Norwegian Government",
            "jurisdiction": "Norway",
            "authoritative_language": "English",
            "source_family": "Public authority",
            "document_type": "Regulation",
            "requirement_role": "Primary normative",
            "inclusion_rationale": "Potentially relevant official regulation.",
        }
        self.inbox.write_text(json.dumps([candidate]), encoding="utf-8")
        human_operations.sync_discovery_candidates(self.config_path)
        workbook = load_workbook(self.workbook)
        self.assertEqual(workbook["Source Register"].max_row, 3)
        workbook.close()
        row = self._operation_rows(operation_type="NEW_SOURCE_CANDIDATE")[0]
        self._set_operation(row, "ACCEPT", note="Identity and scope confirmed.")
        human_operations.process_decisions(self.config_path)
        self.assertEqual(self._source_record("PA002")["source_title"], "New official regulation")

        rejected = dict(candidate, source_title="Out-of-scope notice", official_url="https://www.regjeringen.no/en/out-of-scope/")
        self.inbox.write_text(json.dumps([candidate, rejected]), encoding="utf-8")
        human_operations.sync_discovery_candidates(self.config_path)
        row = self._operation_rows(operation_type="NEW_SOURCE_CANDIDATE")[-1]
        self._set_operation(row, "REJECT", note="Outside the approved source scope.")
        human_operations.process_decisions(self.config_path)
        workbook = load_workbook(self.workbook)
        self.assertEqual(workbook["Source Register"].max_row, 4)
        sheet = workbook["Human Operation Desktop"]
        headers = self._operation_headers(sheet)
        self.assertEqual(sheet.cell(row, headers["program_status"]).value, "REJECTED")
        workbook.close()

    def test_manual_file_intake_creates_candidate_then_preserves_original_html(self):
        url = "https://www.regjeringen.no/en/manual-intake-law/"
        intake_file = self.intake / "manual_intake_law.html"
        intake_file.write_bytes(html_document("Manual intake law", url, "human supplied"))
        result = human_operations.run_human_cycle(self.config_path)
        self.assertEqual(result["manual_intake"]["added_count"], 1)
        self.assertEqual(self._source_record().get("source_id"), "PA001")
        row = self._operation_rows(operation_type="NEW_SOURCE_CANDIDATE")[0]
        self._set_operation(row, "ACCEPT", note="Authorised original HTML verified.")
        human_operations.process_decisions(self.config_path)
        added = self._source_record("PA002")
        self.assertEqual(added["file_format"], "html")
        self.assertEqual(added["current_origin"], "MANUAL")
        self.assertEqual(added["download_status"], "SUCCESS")
        self.assertTrue((self.authority / added["stored_filename"]).is_file())
        self.assertFalse(intake_file.exists())

    def test_manual_replacement_reject_preserves_current_then_accept_archives_it(self):
        rejected_file = self.intake / "replacement_rejected.html"
        rejected_file.write_bytes(html_document("Existing law replacement", self.current_url, "rejected version"))
        human_operations.run_human_cycle(self.config_path)
        row = self._operation_rows(operation_type="MANUAL_FILE_REPLACEMENT")[0]
        self._set_operation(row, "REJECT", note="This copy is not authorised.")
        human_operations.process_decisions(self.config_path)
        self.assertEqual((self.authority / self.current_name).read_bytes(), self.current_bytes)

        workbook = load_workbook(self.workbook)
        source = workbook["Source Register"]
        headers = source_updater.workbook_headers(source, 2)
        source.cell(3, headers["download_status"], "PAYWALL_BLOCKED")
        source.cell(3, headers["failure_stage"], "HTTP")
        source.cell(3, headers["failure_reason"], "Official access requires an authorised copy.")
        workbook.save(self.workbook)
        workbook.close()

        accepted_file = self.intake / "replacement_accepted.html"
        accepted_bytes = html_document("Existing law replacement", self.current_url, "accepted version")
        accepted_file.write_bytes(accepted_bytes)
        human_operations.run_human_cycle(self.config_path)
        row = self._operation_rows(operation_type="MANUAL_FILE_REPLACEMENT")[-1]
        self._set_operation(row, "ACCEPT", note="Authorised replacement confirmed.")
        human_operations.process_decisions(self.config_path)
        record = self._source_record()
        self.assertEqual(record["snapshot_id"], "PA001-002")
        self.assertEqual(record["current_origin"], "MANUAL")
        self.assertEqual(record["download_status"], "SUCCESS")
        self.assertEqual((self.authority / record["stored_filename"]).read_bytes(), accepted_bytes)
        self.assertEqual((self.authority / "_archive" / "PA001" / self.current_name).read_bytes(), self.current_bytes)

    def test_random_qa_correct_and_incorrect_retains_history_and_creates_followup(self):
        first = human_operations.generate_random_qa(self.config_path, random.Random(1), today=date(2026, 8, 31))
        self.assertEqual(first["generated_count"], 1)
        row = self._operation_rows(operation_type="RANDOM_QA_CHECK")[0]
        self._set_operation(row, "CORRECT", note="Source, URL, assessment and file are correct.")
        human_operations.process_decisions(self.config_path)

        second = human_operations.generate_random_qa(self.config_path, random.Random(1), today=date(2026, 9, 30))
        self.assertEqual(second["generated_count"], 1)
        row = self._operation_rows(operation_type="RANDOM_QA_CHECK")[-1]
        self._set_operation(row, "INCORRECT", note="Issuer name must be corrected.")
        human_operations.process_decisions(self.config_path)
        followups = self._operation_rows(trigger="RANDOM_QA_FAILURE")
        self.assertEqual(len(followups), 1)
        workbook = load_workbook(self.workbook)
        sheet = workbook["Human Operation Desktop"]
        headers = self._operation_headers(sheet)
        self.assertEqual(sheet.cell(followups[0], headers["program_status"]).value, "PENDING")
        self.assertTrue(bool(sheet.row_dimensions[row].hidden))
        workbook.close()

    def test_missing_operator_and_stale_fingerprint_are_safely_returned(self):
        review = {
            "review_id": "stale-review",
            "source_id": "PA001",
            "source_title": "Existing law",
            "trigger": "SELECTION_PENDING",
            "reason": "Review selection.",
            "recommended_action": "Apply governed fields.",
        }
        human_operations.sync_review_queue(self.config_path, [review])
        row = self._operation_rows()[0]
        self._set_operation(row, "APPLY", operator="", note="Reviewed.")
        human_operations.process_decisions(self.config_path)
        workbook = load_workbook(self.workbook)
        operations = workbook["Human Operation Desktop"]
        op_headers = self._operation_headers(operations)
        self.assertEqual(operations.cell(row, op_headers["program_status"]).value, "WAITING_FOR_HUMAN")
        source = workbook["Source Register"]
        source_headers = source_updater.workbook_headers(source, 2)
        source.cell(3, source_headers["notes"], "Newer state written after task creation.")
        operations.cell(row, op_headers["decision"], "APPLY")
        operations.cell(row, op_headers["operator"], "Ana")
        workbook.save(self.workbook)
        workbook.close()
        human_operations.process_decisions(self.config_path)
        workbook = load_workbook(self.workbook)
        operations = workbook["Human Operation Desktop"]
        self.assertIn("changed after this task", operations.cell(row, op_headers["program_note"]).value)
        workbook.close()

    def test_manual_intake_is_idempotent_and_excel_open_defers_without_writing(self):
        intake_file = self.intake / "single_candidate.html"
        intake_file.write_bytes(html_document("Single candidate", "https://www.regjeringen.no/en/single/"))
        first = manual_intake.scan_manual_intake(self.config_path)
        second = manual_intake.scan_manual_intake(self.config_path)
        self.assertEqual(first["added_count"], 1)
        self.assertEqual(second["added_count"], 0)

        before = hashlib.sha256(self.workbook.read_bytes()).hexdigest()
        lock = self.workbook.with_name(f"~${self.workbook.name}")
        lock.write_text("Excel lock simulation", encoding="utf-8")
        code, result = run_cycle(self.config_path, full=False)
        self.assertEqual(code, 2)
        self.assertEqual(result["status"], "DEFERRED_WORKBOOK_OPEN")
        self.assertEqual(hashlib.sha256(self.workbook.read_bytes()).hexdigest(), before)

    def test_no_api_routine_cycle_is_deterministic_and_never_calls_downloader(self):
        with patch("source_updater.run_updates") as downloader, redirect_stdout(io.StringIO()):
            code, report = leader_orchestrator.run_leader(
                self.config_path, execute=False, include_pending=True, force_all=True,
            )
        self.assertEqual(code, 0)
        downloader.assert_not_called()
        self.assertEqual(report["capability_mode"], "NO_API")
        self.assertEqual(report["provider_extensions"]["status"], "DISABLED")
        collection = next(
            item for item in report["module_results"] if item["module"] == "retrieval_and_monitoring"
        )["collection"]
        self.assertFalse(collection["executed"])

    def test_simulated_api_proposals_are_human_gated_before_any_register_change(self):
        class FakeDiscovery:
            name = "Fake Search API"

            def discover(self, records):
                return [{
                    "source_title": "API-proposed regulation",
                    "official_url": "https://www.regjeringen.no/en/api-proposed/",
                    "retrieval_url": "https://www.regjeringen.no/en/api-proposed/",
                    "folder_code": "A_Public_Authority",
                    "file_format": "html",
                    "issuer": "Norwegian Government",
                    "jurisdiction": "Norway",
                    "authoritative_language": "English",
                    "source_family": "Public authority",
                    "document_type": "Regulation",
                    "requirement_role": "Primary normative",
                    "inclusion_rationale": "API draft requiring human confirmation.",
                }]

        class FakeAssessment:
            name = "Fake Assessment API"

            def assess(self, records):
                return [{
                    "source_id": "PA001",
                    "reason": "Scope score needs a human re-check.",
                    "recommended_action": "Confirm the proposed MEDIUM scope score.",
                    "proposed_updates": {"scope_relevance": "MEDIUM"},
                }]

        with redirect_stdout(io.StringIO()):
            code, report = leader_orchestrator.run_leader(
                self.config_path,
                execute=False,
                include_pending=True,
                force_all=True,
                providers={"discovery": FakeDiscovery(), "assessment": FakeAssessment()},
            )
        self.assertEqual(code, 0)
        self.assertEqual(report["capability_mode"], "API_CONNECTED")
        self.assertEqual(report["provider_extensions"]["status"], "PASS")
        workbook = load_workbook(self.workbook)
        self.assertEqual(workbook["Source Register"].max_row, 3)
        workbook.close()
        self.assertEqual(self._source_record()["scope_relevance"], "HIGH")

        candidate_row = self._operation_rows(operation_type="NEW_SOURCE_CANDIDATE")[0]
        review_row = self._operation_rows(trigger="API_ASSESSMENT_SUGGESTION")[0]
        self._set_operation(candidate_row, "ACCEPT", note="API candidate verified by a human.")
        self._set_operation(review_row, "APPLY", note="MEDIUM scope score confirmed by a human.")
        human_operations.process_decisions(self.config_path)
        self.assertEqual(self._source_record("PA002")["source_title"], "API-proposed regulation")
        self.assertEqual(self._source_record()["scope_relevance"], "MEDIUM")

    def test_simulated_api_failure_degrades_extension_but_core_cycle_completes(self):
        class FailingDiscovery:
            name = "Unavailable Search API"

            def discover(self, records):
                raise TimeoutError("simulated timeout")

        class UnsafeAssessment:
            name = "Unsafe Assessment API"

            def assess(self, records):
                return [{
                    "source_id": "PA001",
                    "reason": "Attempt to change a program-owned field.",
                    "proposed_updates": {"content_hash": "not-allowed"},
                }]

        with redirect_stdout(io.StringIO()):
            code, report = leader_orchestrator.run_leader(
                self.config_path,
                execute=False,
                include_pending=True,
                force_all=True,
                providers={"discovery": FailingDiscovery(), "assessment": UnsafeAssessment()},
            )
        self.assertEqual(code, 0)
        self.assertEqual(report["overall_status"], "REVIEW")
        self.assertEqual(report["provider_extensions"]["status"], "DEGRADED")
        self.assertEqual(len(report["provider_extensions"]["errors"]), 2)
        self.assertEqual(self._source_record()["content_hash"], hashlib.sha256(self.current_bytes).hexdigest())
        self.assertIn("deterministic cycle completed", report["user_message"])


if __name__ == "__main__":
    unittest.main()
