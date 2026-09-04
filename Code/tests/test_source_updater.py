from __future__ import annotations

import hashlib
import io
import json
import sys
import tempfile
import threading
import unittest
import zipfile
from datetime import date
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import source_updater as updater


HEADERS = [
    "source_id", "snapshot_id", "source_title", "official_url", "retrieval_url",
    "folder_code", "file_format", "stored_filename", "snapshot_status", "issuer",
    "jurisdiction", "authoritative_language", "version", "effective_date", "source_family",
    "document_type", "requirement_role", "inclusion_rationale", "authority_quality", "scope_relevance",
    "version_currency", "traceability", "access_permission", "automation_readiness", "selection_status",
    "download_status", "last_attempt_at", "last_success_at", "content_change", "current_origin",
    "current_snapshot_date", "manual_update_date", "manual_updated_by", "failure_stage", "failure_reason",
    "needs_human_action", "content_hash", "source_status", "primary_source_id", "notes",
    "acquisition_channel", "provenance_status", "update_model", "applicability_reference",
    "operator_selection_decision",
]


def pdf_payload(label: str) -> bytes:
    return b"%PDF-1.4\n" + label.encode("utf-8") + b"\n" + (b"x" * 700) + b"\n%%EOF\n"


def zip_payload(xlsx: bool = False) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("payload.txt", "test-source-package" * 20)
        if xlsx:
            archive.writestr("[Content_Types].xml", "<Types></Types>")
            archive.writestr("xl/workbook.xml", "<workbook></workbook>")
    return buffer.getvalue()


class MutableHandler(BaseHTTPRequestHandler):
    routes = {}
    last_accept = None

    def do_GET(self):
        type(self).last_accept = self.headers.get("Accept")
        status, content_type, payload = self.routes.get(
            self.path, (404, "text/plain", b"not found")
        )
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def log_message(self, *args):
        return


class SourceUpdaterTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.library = self.root / "library"
        self.updater_dir = self.library / "Updater"
        self.updater_dir.mkdir(parents=True)
        self.source_root = self.library / "Data"
        (self.source_root / "C_Certification_Scheme").mkdir(parents=True)
        self.workbook = self.library / "Requirement_Source_Registry.xlsx"
        self.server = ThreadingHTTPServer(("127.0.0.1", 0), MutableHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.server.server_address[1]}"
        MutableHandler.routes = {
            "/source.pdf": (200, "application/pdf", pdf_payload("version-one")),
            "/source.html": (200, "application/xhtml+xml", b"<!doctype html><html><head><title>Official source</title></head><body>" + (b"English legal text " * 20) + b"</body></html>"),
            "/manual.pdf": (200, "application/pdf", pdf_payload("manual-version")),
            "/login.html": (200, "text/html", b"<!doctype html><html><head><title>Login</title></head><body>Sign in</body></html>"),
            "/source.xlsx": (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", zip_payload(xlsx=True)),
            "/source.zip": (200, "application/zip", zip_payload()),
        }
        self._make_workbook(f"{self.base_url}/source.pdf", "pdf")
        self.config_path = self.updater_dir / "config.json"
        self.config_path.write_text(json.dumps({
            "workbook": "../Requirement_Source_Registry.xlsx",
            "source_root": "../Data",
            "backup_root": "../_backups",
            "log_root": "../Logs",
            "timezone": "Europe/Oslo",
            "minimum_pdf_bytes": 100,
            "minimum_html_bytes": 50,
            "minimum_binary_bytes": 50,
            "timeout_seconds": 5,
        }), encoding="utf-8")
        self.config = updater.read_config(self.config_path)

    def tearDown(self):
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=2)
        self.temp.cleanup()

    def _make_workbook(self, retrieval_url: str, file_format: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Source Register"
        for column, header in enumerate(HEADERS, 1):
            ws.cell(row=2, column=column, value=header)
        values = {
            "source_id": "CS001",
            "snapshot_id": "CS001-001",
            "source_title": "Test Source",
            "official_url": retrieval_url,
            "retrieval_url": retrieval_url,
            "folder_code": "C_Certification_Scheme",
            "file_format": file_format,
            "stored_filename": f"CS001-001_Test_Source.{file_format}",
            "snapshot_status": "NOT_COLLECTED",
            "issuer": "Test Issuer",
            "source_family": "Certification scheme",
            "requirement_role": "Primary normative",
            "authority_quality": "HIGH",
            "scope_relevance": "HIGH",
            "version_currency": "HIGH",
            "traceability": "HIGH",
            "access_permission": "HIGH",
            "automation_readiness": "HIGH",
            "operator_selection_decision": "INCLUDE",
            "acquisition_channel": "OFFICIAL_WEBSITE",
            "provenance_status": "OFFICIAL_ORIGINAL",
            "update_model": "VERSIONED",
            "selection_status": "=IF(COUNTIF(S3:W3,\"HIGH\")=5,\"INCLUDE\",\"PENDING\")",
            "download_status": "NOT_RUN",
            "content_change": "UNKNOWN",
            "current_origin": "NONE",
            "current_snapshot_date": "=IF(AD3=\"AUTO\",AB3,AF3)",
            "failure_stage": "NONE",
            "needs_human_action": "=IF(Z3=\"FAIL\",\"YES\",\"NO\")",
            "source_status": "CURRENT",
        }
        for column, header in enumerate(HEADERS, 1):
            ws.cell(row=3, column=column, value=values.get(header))
        wb.save(self.workbook)

    def _row(self):
        wb = load_workbook(self.workbook, data_only=False)
        ws = wb["Source Register"]
        headers = {cell.value: cell.column for cell in ws[2] if cell.value}
        return wb, ws, headers, updater.record_from_row(ws, 3, headers)

    def test_success_then_unchanged(self):
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, first = self._row()
        self.assertEqual(first["download_status"], "SUCCESS")
        self.assertEqual(first["content_change"], "INITIAL")
        self.assertEqual(first["snapshot_id"], "CS001-001")
        self.assertEqual(first["current_origin"], "AUTO")
        current = self.source_root / first["folder_code"] / first["stored_filename"]
        self.assertTrue(current.is_file())
        self.assertEqual(first["content_hash"], hashlib.sha256(current.read_bytes()).hexdigest())

        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, second = self._row()
        self.assertEqual(second["content_change"], "UNCHANGED")
        self.assertEqual(second["snapshot_id"], "CS001-001")

    def test_changed_archives_old_snapshot(self):
        updater.run_updates(self.config, set(), False)
        MutableHandler.routes["/source.pdf"] = (200, "application/pdf", pdf_payload("version-two"))
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, record = self._row()
        self.assertEqual(record["snapshot_id"], "CS001-002")
        self.assertEqual(record["content_change"], "CHANGED")
        archive = self.source_root / "C_Certification_Scheme" / "_archive" / "CS001" / "CS001-001_Test_Source.pdf"
        self.assertTrue(archive.is_file())
        self.assertFalse((self.source_root / "C_Certification_Scheme" / "CS001-001_Test_Source.pdf").exists())

    def test_http_failure_preserves_current(self):
        updater.run_updates(self.config, set(), False)
        _, _, _, before = self._row()
        current = self.source_root / before["folder_code"] / before["stored_filename"]
        before_bytes = current.read_bytes()
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["retrieval_url"], f"{self.base_url}/missing.pdf")
        wb.save(self.workbook)

        self.assertEqual(updater.run_updates(self.config, set(), False), 1)
        _, _, _, after = self._row()
        self.assertEqual(after["download_status"], "FAIL")
        self.assertEqual(after["failure_stage"], "HTTP")
        self.assertEqual(current.read_bytes(), before_bytes)
        self.assertEqual(after["snapshot_id"], before["snapshot_id"])

    def test_manual_fallback_survives_failure_then_auto_takes_over(self):
        updater.run_updates(self.config, set(), False)
        wb, ws, headers, record = self._row()
        current = self.source_root / record["folder_code"] / record["stored_filename"]
        manual = pdf_payload("manual-version")
        current.write_bytes(manual)
        ws.cell(3, headers["retrieval_url"], f"{self.base_url}/missing.pdf")
        ws.cell(3, headers["current_origin"], "MANUAL")
        ws.cell(3, headers["manual_update_date"], date.today())
        ws.cell(3, headers["manual_updated_by"], "Tester")
        ws.cell(3, headers["content_hash"], hashlib.sha256(manual).hexdigest())
        wb.save(self.workbook)

        self.assertEqual(updater.run_updates(self.config, set(), False), 1)
        _, _, _, failed = self._row()
        self.assertEqual(failed["current_origin"], "MANUAL")
        self.assertEqual(current.read_bytes(), manual)

        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["retrieval_url"], f"{self.base_url}/manual.pdf")
        wb.save(self.workbook)
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, recovered = self._row()
        self.assertEqual(recovered["download_status"], "SUCCESS")
        self.assertEqual(recovered["content_change"], "UNCHANGED")
        self.assertEqual(recovered["current_origin"], "AUTO")
        self.assertEqual(current.read_bytes(), manual)

    def test_html_login_page_fails_content_check(self):
        self._make_workbook(f"{self.base_url}/login.html", "html")
        self.assertEqual(updater.run_updates(self.config, set(), False), 1)
        _, _, _, record = self._row()
        self.assertEqual(record["download_status"], "FAIL")
        self.assertEqual(record["failure_stage"], "CONTENT")

    def test_html_request_accepts_xhtml_content_negotiation(self):
        self._make_workbook(f"{self.base_url}/source.html", "html")
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        self.assertIn("application/xhtml+xml", MutableHandler.last_accept)
        _, _, _, record = self._row()
        self.assertEqual(record["download_status"], "SUCCESS")

    def test_xlsx_original_format_is_validated_and_stored(self):
        self._make_workbook(f"{self.base_url}/source.xlsx", "xlsx")
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, record = self._row()
        self.assertEqual(record["download_status"], "SUCCESS")
        self.assertTrue((self.source_root / record["folder_code"] / record["stored_filename"]).is_file())

    def test_zip_original_format_is_validated_and_stored(self):
        self._make_workbook(f"{self.base_url}/source.zip", "zip")
        self.assertEqual(updater.run_updates(self.config, set(), False), 0)
        _, _, _, record = self._row()
        self.assertEqual(record["download_status"], "SUCCESS")
        self.assertTrue((self.source_root / record["folder_code"] / record["stored_filename"]).is_file())

    def test_available_only_skips_missing_url_without_recording_failure(self):
        self._make_workbook("", "pdf")
        self.assertEqual(updater.run_updates(self.config, set(), False, available_only=True), 0)
        _, _, _, record = self._row()
        self.assertEqual(record["download_status"], "NOT_RUN")
        self.assertEqual(record["snapshot_status"], "NOT_COLLECTED")
        self.assertEqual(record["failure_stage"], "NONE")

    def test_unverified_copy_remains_pending(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["provenance_status"], "UNVERIFIED")
        wb.save(self.workbook)
        _, _, _, record = self._row()
        self.assertEqual(updater.selection_from_scores(record), "PENDING")

    def test_official_channel_requires_official_url(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["official_url"], "")
        wb.save(self.workbook)
        _, _, _, record = self._row()
        self.assertEqual(updater.selection_from_scores(record), "PENDING")

    def test_private_channel_collection_is_independent_of_selection(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["official_url"], "")
        ws.cell(3, headers["acquisition_channel"], "CLIENT_OR_SITE")
        ws.cell(3, headers["provenance_status"], "AUTHORISED_COPY")
        ws.cell(3, headers["automation_readiness"], "MEDIUM")
        wb.save(self.workbook)
        _, _, _, record = self._row()
        self.assertEqual(updater.selection_from_scores(record), "PENDING")
        self.assertTrue(updater.eligible(record, self.config, False))

    def test_equipment_manual_requires_manufacturer_and_applicability(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["document_type"], "Equipment manual / OEM instruction")
        ws.cell(3, headers["source_family"], "Manufacturer / supplier")
        ws.cell(3, headers["inclusion_rationale"], "Applicable operating limits are mandatory for the installed asset.")
        wb.save(self.workbook)
        _, _, _, missing = self._row()
        self.assertEqual(updater.selection_from_scores(missing), "PENDING")

        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["applicability_reference"], "Site A / Model X / Asset 42")
        wb.save(self.workbook)
        _, _, _, complete = self._row()
        self.assertEqual(updater.selection_from_scores(complete), "PENDING")
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["download_status"], "SUCCESS")
        ws.cell(3, headers["snapshot_status"], "STORED")
        wb.save(self.workbook)
        _, _, _, collected = self._row()
        self.assertEqual(updater.selection_from_scores(collected), "INCLUDE")

    def test_audit_detects_missing_stored_file(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["snapshot_status"], "STORED")
        wb.save(self.workbook)
        issues = updater.audit_registry(self.config)
        self.assertTrue(any("file does not exist" in issue for issue in issues))

    def test_audit_detects_source_prefix_folder_mismatch(self):
        wb, ws, headers, _ = self._row()
        ws.cell(3, headers["source_id"], "PA001")
        ws.cell(3, headers["snapshot_id"], "PA001-001")
        ws.cell(3, headers["stored_filename"], "PA001-001_Test_Source.pdf")
        wb.save(self.workbook)
        issues = updater.audit_registry(self.config)
        self.assertTrue(any("prefix must be CS" in issue for issue in issues))


if __name__ == "__main__":
    unittest.main()
