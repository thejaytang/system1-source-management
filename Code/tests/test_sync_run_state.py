from __future__ import annotations

import json
import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import source_updater as updater
import sync_run_state as syncer


HEADERS = [
    "source_id", "snapshot_id", "source_title", "official_url", "retrieval_url", "folder_code",
    "file_format", "stored_filename", "snapshot_status", "issuer", "jurisdiction", "authoritative_language",
    "version", "effective_date", "source_family", "document_type", "requirement_role", "inclusion_rationale",
    "authority_quality", "scope_relevance", "version_currency", "traceability", "access_permission",
    "automation_readiness", "selection_status", "download_status", "last_attempt_at", "last_success_at",
    "content_change", "current_origin", "current_snapshot_date", "manual_update_date", "manual_updated_by",
    "failure_stage", "failure_reason", "needs_human_action", "content_hash", "source_status",
    "primary_source_id", "notes", "acquisition_channel", "provenance_status", "update_model", "applicability_reference",
    "operator_selection_decision",
]


def make_workbook(path: Path, *, title: str, origin: str, status: str, manual_date=None, attempt=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Source Register"
    for column, header in enumerate(HEADERS, 1):
        ws.cell(2, column, header)
    values = {
        "source_id": "PA001", "snapshot_id": "PA001-001", "source_title": title,
        "official_url": "https://example.test/official", "retrieval_url": "https://example.test/source",
        "folder_code": "A_Public_Authority", "file_format": "html",
        "stored_filename": "PA001-001_Test.html", "snapshot_status": "STORED" if status == "SUCCESS" else "NOT_COLLECTED",
        "issuer": "Test Authority", "source_family": "Public authority", "requirement_role": "Primary normative",
        "inclusion_rationale": "Team-owned rationale", "authority_quality": "HIGH", "scope_relevance": "HIGH",
        "version_currency": "HIGH", "traceability": "HIGH", "access_permission": "HIGH",
        "automation_readiness": "HIGH", "operator_selection_decision": "INCLUDE",
        "selection_status": "INCLUDE", "download_status": status,
        "last_attempt_at": attempt, "last_success_at": attempt if status == "SUCCESS" else None,
        "content_change": "CHANGED" if status == "SUCCESS" else "UNKNOWN", "current_origin": origin,
        "manual_update_date": manual_date, "manual_updated_by": "Human" if origin == "MANUAL" else "",
        "failure_stage": "NONE", "failure_reason": "", "content_hash": "abc" if status == "SUCCESS" else "",
        "source_status": "CURRENT", "notes": "Team-owned note", "acquisition_channel": "OFFICIAL_WEBSITE",
        "provenance_status": "OFFICIAL_ORIGINAL", "update_model": "ROLLING", "applicability_reference": "",
    }
    for column, header in enumerate(HEADERS, 1):
        ws.cell(3, column, values.get(header))
    wb.save(path)


class SyncRunStateTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.library = self.root / "library"
        self.updater_dir = self.library / "Updater"
        self.updater_dir.mkdir(parents=True)
        self.live = self.library / "Requirement_Source_Registry.xlsx"
        self.staged = self.root / "staged.xlsx"
        self.config_path = self.updater_dir / "config.json"
        self.config_path.write_text(json.dumps({
            "workbook": "../Requirement_Source_Registry.xlsx",
            "source_root": "../Data",
            "backup_root": "../_backups",
            "log_root": "../Logs",
        }), encoding="utf-8")
        self.config = updater.read_config(self.config_path)

    def tearDown(self):
        self.temp.cleanup()

    def test_sync_updates_only_program_fields(self):
        make_workbook(self.live, title="Human title", origin="NONE", status="NOT_RUN")
        make_workbook(self.staged, title="Human title", origin="AUTO", status="SUCCESS", attempt=datetime(2026, 8, 28, 10, 0))
        result = syncer.sync_state(self.config, self.staged)
        self.assertEqual(result["synced"], ["PA001"])
        wb = load_workbook(self.live, data_only=False)
        ws = wb["Source Register"]
        headers = {cell.value: cell.column for cell in ws[2] if cell.value}
        self.assertEqual(ws.cell(3, headers["download_status"]).value, "SUCCESS")
        self.assertEqual(ws.cell(3, headers["current_origin"]).value, "AUTO")
        self.assertEqual(ws.cell(3, headers["source_title"]).value, "Human title")
        self.assertEqual(ws.cell(3, headers["notes"]).value, "Team-owned note")

    def test_newer_manual_snapshot_wins(self):
        make_workbook(self.live, title="Human title", origin="MANUAL", status="SUCCESS", manual_date=date(2026, 8, 29))
        make_workbook(self.staged, title="Human title", origin="AUTO", status="SUCCESS", attempt=datetime(2026, 8, 28, 10, 0))
        result = syncer.sync_state(self.config, self.staged)
        self.assertEqual(result["synced"], [])
        self.assertEqual(result["conflicts"][0]["reason"], "newer_manual_snapshot_preserved")
        wb = load_workbook(self.live, data_only=False)
        ws = wb["Source Register"]
        headers = {cell.value: cell.column for cell in ws[2] if cell.value}
        self.assertEqual(ws.cell(3, headers["current_origin"]).value, "MANUAL")


if __name__ == "__main__":
    unittest.main()
