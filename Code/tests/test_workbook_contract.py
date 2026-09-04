from __future__ import annotations

import json
import shutil
import sys
import tempfile
import threading
import unittest
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import source_updater as updater


def pdf_payload() -> bytes:
    return b"%PDF-1.4\ncontract-test\n" + (b"x" * 700) + b"\n%%EOF\n"


class PdfHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        payload = pdf_payload()
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def log_message(self, *args):
        return


class FinalWorkbookContractTest(unittest.TestCase):
    def test_dashboard_management_reporting_contract(self):
        final_workbook = Path(__file__).resolve().parents[2] / "Requirement_Source_Registry.xlsx"
        workbook = load_workbook(final_workbook, data_only=False)
        dashboard = workbook["Dashboard"]

        self.assertEqual(dashboard["A1"].value, "SYSTEM1 SOURCE MANAGEMENT")
        self.assertIn("Read-only management overview", dashboard["A2"].value)
        self.assertEqual(
            [dashboard[cell].value for cell in ("A4", "C4", "E4", "H4", "L4")],
            ["TOTAL SOURCES", "INCLUDED SOURCES", "STORED SNAPSHOTS", "OPEN HUMAN TASKS", "RETRIEVAL EXCEPTIONS"],
        )
        for cell in ("A5", "C5", "E5", "H5", "L5"):
            self.assertIn("Table[", dashboard[cell].value)
        self.assertIn("Human review is overdue", dashboard["A9"].value)
        self.assertEqual([dashboard.cell(35, column).value for column in range(1, 7)], [
            "SOURCE ID", "SOURCE TITLE", "TASK TYPE", "ISSUE SUMMARY", "DAYS OPEN", "PROGRAM STATUS",
        ])
        self.assertIn("Complete decisions in Human Operation Desktop", dashboard["A42"].value)
        self.assertEqual(len(dashboard._charts), 2)
        self.assertTrue(any(getattr(chart, "type", None) == "bar" for chart in dashboard._charts))
        self.assertTrue(any(chart.__class__.__name__ == "DoughnutChart" for chart in dashboard._charts))
        hidden_ranges = [
            (dimension.min, dimension.max)
            for dimension in dashboard.column_dimensions.values()
            if dimension.hidden
        ]
        for column_index in range(18, 23):
            self.assertTrue(any(start <= column_index <= end for start, end in hidden_ranges))
        self.assertFalse(dashboard.sheet_view.showGridLines)
        self.assertEqual(dashboard.sheet_view.zoomScale, 85)

        values_workbook = load_workbook(final_workbook, data_only=True)
        register = values_workbook["Source Register"]
        register_headers = {cell.value: cell.column for cell in register[2] if cell.value}
        register_rows = [
            row for row in range(3, register.max_row + 1)
            if register.cell(row, register_headers["source_id"]).value
        ]
        operations = values_workbook["Human Operation Desktop"]
        operation_headers = {cell.value: cell.column for cell in operations[2] if cell.value}
        operation_rows = [
            row for row in range(3, operations.max_row + 1)
            if operations.cell(row, operation_headers["source_id"]).value
        ]
        self.assertEqual(len(register_rows), 87)
        self.assertEqual(sum(
            register.cell(row, register_headers["selection_status"]).value == "INCLUDE"
            for row in register_rows
        ), 44)
        self.assertEqual(sum(
            register.cell(row, register_headers["snapshot_status"]).value == "STORED"
            for row in register_rows
        ), 70)
        self.assertEqual(sum(
            operations.cell(row, operation_headers["program_status"]).value
            in {"PENDING", "WAITING_FOR_HUMAN", "NEEDS_REPLAN"}
            for row in operation_rows
        ), 49)
        self.assertEqual(sum(
            register.cell(row, register_headers["download_status"]).value
            in {"FAIL", "PAYWALL_BLOCKED"}
            for row in register_rows
        ), 8)
        values_workbook.close()
        workbook.close()

    def test_updater_preserves_english_contract_dashboard_and_excel_features(self):
        final_workbook = Path(__file__).resolve().parents[2] / "Requirement_Source_Registry.xlsx"
        self.assertTrue(final_workbook.is_file())
        with tempfile.TemporaryDirectory() as temp_name:
            library = Path(temp_name) / "Requirement Source Management"
            updater_dir = library / "Code" / "config"
            updater_dir.mkdir(parents=True)
            workbook = library / final_workbook.name
            shutil.copy2(final_workbook, workbook)
            source_root = library / "Data"

            server = ThreadingHTTPServer(("127.0.0.1", 0), PdfHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                url = f"http://127.0.0.1:{server.server_address[1]}/source.pdf"
                wb = load_workbook(workbook, data_only=False)
                ws = wb["Source Register"]
                headers = {cell.value: cell.column for cell in ws[2] if cell.value}
                source_id = ws.cell(3, headers["source_id"]).value
                folder_code = ws.cell(3, headers["folder_code"]).value
                (source_root / folder_code).mkdir(parents=True)
                ws.cell(3, headers["retrieval_url"], url)
                ws.cell(3, headers["file_format"], "pdf")
                stored_filename = str(ws.cell(3, headers["stored_filename"]).value)
                ws.cell(3, headers["stored_filename"], str(Path(stored_filename).with_suffix(".pdf")))
                for field in updater.SCORE_FIELDS:
                    ws.cell(3, headers[field], "HIGH")
                wb.save(workbook)

                before = load_workbook(workbook, data_only=False)
                before_ws = before["Source Register"]
                before_dashboard = before["Dashboard"]
                feature_counts = (
                    len(before_ws.data_validations.dataValidation),
                    len(before_ws.conditional_formatting),
                    len(before_ws.tables),
                    len(before_ws._charts),
                    len(before_dashboard._charts),
                )
                formulas = {
                    "selection": before_ws["E3"].value,
                    "snapshot_date": before_ws["K3"].value,
                    "human_action": before_ws["N3"].value,
                    "dashboard_total": before_dashboard["A4"].value,
                }

                config_path = updater_dir / "config.json"
                config_path.write_text(json.dumps({
                    "workbook": "../../Requirement_Source_Registry.xlsx",
                    "source_root": "../../Data",
                    "backup_root": "../runtime/backups",
                    "log_root": "../runtime/logs",
                    "timezone": "Europe/Oslo",
                    "minimum_pdf_bytes": 100,
                    "timeout_seconds": 5,
                }), encoding="utf-8")
                config = updater.read_config(config_path)
                result = updater.run_updates(config, {source_id}, False)
                if result != 0:
                    failed = load_workbook(workbook, data_only=False)["Source Register"]
                    stage = failed.cell(3, headers["failure_stage"]).value
                    reason = failed.cell(3, headers["failure_reason"]).value
                    self.fail(f"Updater contract run failed at {stage}: {reason}")

                after = load_workbook(workbook, data_only=False)
                after_ws = after["Source Register"]
                after_dashboard = after["Dashboard"]
                self.assertEqual(after.sheetnames, ["Instructions", "Categories", "Source Register", "Dashboard", "Human Operation Desktop"])
                self.assertEqual(
                    (
                        len(after_ws.data_validations.dataValidation),
                        len(after_ws.conditional_formatting),
                        len(after_ws.tables),
                        len(after_ws._charts),
                        len(after_dashboard._charts),
                    ),
                    feature_counts,
                )
                self.assertEqual(after_ws["E3"].value, formulas["selection"])
                self.assertEqual(after_ws["K3"].value, formulas["snapshot_date"])
                self.assertEqual(after_ws["N3"].value, formulas["human_action"])
                self.assertEqual(after_dashboard["A4"].value, formulas["dashboard_total"])
                self.assertEqual(after_ws.cell(3, headers["download_status"]).value, "SUCCESS")
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2)


if __name__ == "__main__":
    unittest.main()
